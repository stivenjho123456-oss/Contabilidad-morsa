import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app_paths import default_writable_excel_path
from backup_manager import create_backup_if_due
from database import (
    clear_nomina,
    init_db,
    save_nomina_asistencia,
    save_nomina_resumen,
    save_nomina_seg_social,
    write_transaction,
)


EXCEL_NOMINA_FILENAME = 'NOMINA DE  FEBRERO -2026.xlsx'
EXCEL_NOMINA_PATH = str(default_writable_excel_path(EXCEL_NOMINA_FILENAME))


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def as_float(value):
    if value in (None, ''):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        raw = str(value).strip().replace('$', '').replace('.', '').replace(',', '.')
        try:
            return float(raw)
        except ValueError:
            return 0.0


def detect_periodo(wb):
    ws = wb['NOM_FIJOS']
    raw = clean_text(ws['C6'].value)
    match = re.search(r'([A-ZÁÉÍÓÚÑ]+)\s+DE\s+(\d{4})', raw.upper())
    if not match:
        return 'NOMINA'
    return f'{match.group(1)} {match.group(2)}'


def parse_desprendibles(ws, periodo, origen_archivo, conn):
    count = 0
    for row in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row, 2).value).upper() != 'LA MORSA FOODS SAS':
            continue

        empleado = clean_text(ws.cell(row + 2, 4).value)
        if not empleado:
            continue

        q1 = {
            'dias': 0.0,
            'devengado': 0.0,
            'aux': 0.0,
            'salud': 0.0,
            'pension': 0.0,
            'neto': 0.0,
        }
        q2 = {
            'dias': 0.0,
            'devengado': 0.0,
            'aux': 0.0,
            'salud': 0.0,
            'pension': 0.0,
            'neto': 0.0,
        }
        resumen_mes = {}
        for offset in range(4, 11):
            label_q1 = clean_text(ws.cell(row + offset, 2).value).upper()
            sub_q1 = clean_text(ws.cell(row + offset, 3).value).upper()
            value_q1 = as_float(ws.cell(row + offset, 5).value)

            if label_q1 == 'VALOR DIA':
                q1['dias'] = value_q1
            elif label_q1 == 'TOTAL QUINCENA':
                q1['devengado'] = value_q1
            elif label_q1 == 'AUXILIO DE TRANSPORTE':
                q1['aux'] = value_q1
            elif label_q1 == 'DEDUCCIONES' and sub_q1 == 'SALUD':
                q1['salud'] = value_q1
            elif sub_q1.startswith('PENS'):
                q1['pension'] = value_q1
            elif label_q1 == 'NETO PAGADO':
                q1['neto'] = value_q1

            label_q2 = clean_text(ws.cell(row + offset, 7).value).upper()
            sub_q2 = clean_text(ws.cell(row + offset, 8).value).upper()
            value_q2 = as_float(ws.cell(row + offset, 10).value)

            if label_q2 == 'VALOR DIA':
                q2['dias'] = value_q2
            elif label_q2 == 'TOTAL QUINCENA':
                q2['devengado'] = value_q2
            elif label_q2 == 'AUXILIO DE TRANSPORTE':
                q2['aux'] = value_q2
            elif label_q2 == 'DEDUCCIONES' and sub_q2 == 'SALUD':
                q2['salud'] = value_q2
            elif sub_q2.startswith('PENS'):
                q2['pension'] = value_q2
            elif label_q2 == 'NETO PAGADO':
                q2['neto'] = value_q2

        for offset in range(5, 11):
            label = clean_text(ws.cell(row + offset, 12).value).upper()
            value = as_float(ws.cell(row + offset, 13).value)
            if label:
                resumen_mes[label] = value

        data = {
            'periodo': periodo,
            'empleado': empleado,
            'cedula': clean_text(ws.cell(row + 3, 4).value),
            'valor_dia': as_float(ws.cell(row + 4, 3).value),
            'q1_dias': q1['dias'],
            'q1_devengado': q1['devengado'],
            'q1_aux_transporte': q1['aux'],
            'q1_salud': q1['salud'],
            'q1_pension': q1['pension'],
            'q1_neto': q1['neto'],
            'q2_dias': q2['dias'],
            'q2_devengado': q2['devengado'],
            'q2_aux_transporte': q2['aux'],
            'q2_salud': q2['salud'],
            'q2_pension': q2['pension'],
            'q2_neto': q2['neto'],
            'total_deduccion': resumen_mes.get(
                'DEDUCCION',
                q1['salud'] + q1['pension'] + q2['salud'] + q2['pension'],
            ),
            'total_incapacidad': resumen_mes.get('INCAPACIDAD', 0.0),
            'total_descuento': resumen_mes.get('DESCUENTO', 0.0),
            'total_mes': resumen_mes.get('TOTAL MES', q1['neto'] + q2['neto']),
            'origen_archivo': origen_archivo,
        }
        save_nomina_resumen(data, conn=conn)
        count += 1
    return count


def parse_seg_social(ws, periodo, origen_archivo, conn):
    count = 0
    grupo_actual = ''
    for row in range(1, ws.max_row + 1):
        concepto = clean_text(ws.cell(row, 3).value)
        valor = ws.cell(row, 4).value
        observacion = clean_text(ws.cell(row, 5).value)

        if observacion:
            grupo_actual = observacion.upper()

        if not concepto or valor in (None, ''):
            continue

        save_nomina_seg_social({
            'periodo': periodo,
            'grupo': grupo_actual,
            'concepto': concepto.upper(),
            'valor': as_float(valor),
            'observaciones': observacion,
            'origen_archivo': origen_archivo,
        }, conn=conn)
        count += 1
    return count


def parse_asistencia(ws, periodo, origen_archivo, conn):
    count = 0
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    first_days = {}
    second_days = {}
    for col, value in enumerate(headers, start=1):
        if isinstance(value, int):
            if 1 <= value <= 15:
                first_days[col] = value
            elif 16 <= value <= 31:
                second_days[col] = value

    for row in range(3, ws.max_row + 1):
        empleado = clean_text(ws.cell(row, 2).value)
        cedula = clean_text(ws.cell(row, 3).value)
        if not empleado or not cedula:
            continue
        for col, dia in {**first_days, **second_days}.items():
            value = ws.cell(row, col).value
            if value in (None, '', 0):
                continue
            try:
                marcado = float(value)
            except (TypeError, ValueError):
                marcado = 0
            if marcado <= 0:
                continue
            save_nomina_asistencia({
                'periodo': periodo,
                'empleado': empleado,
                'cedula': cedula,
                'dia': dia,
                'quincena': 'Q1' if dia <= 15 else 'Q2',
                'estado': 'LABORADO',
                'origen_archivo': origen_archivo,
            }, conn=conn, log_audit=False)
            count += 1
    return count


def migrate_nomina(path=EXCEL_NOMINA_PATH):
    if not os.path.exists(path):
        print(f'ERROR: No se encontro el archivo de nomina:\n  {path}')
        sys.exit(1)

    init_db()
    create_backup_if_due('pre_nomina_import', max_age_minutes=15)
    wb = openpyxl.load_workbook(path, data_only=True)
    periodo = detect_periodo(wb)
    origen_archivo = os.path.basename(path)

    with write_transaction(checkpoint=True, checkpoint_mode='TRUNCATE') as conn:
        clear_nomina(origen_archivo=origen_archivo, conn=conn)
        count_asistencia = parse_asistencia(wb['Asistencia'], periodo, origen_archivo, conn)
        count_nomina = parse_desprendibles(wb['Desprendible FIJOS'], periodo, origen_archivo, conn)
        count_seg = parse_seg_social(wb['SEG SOCIAL'], periodo, origen_archivo, conn)

    print(f'Periodo detectado: {periodo}')
    print(f'Registros de asistencia importados: {count_asistencia}')
    print(f'Registros de nomina importados: {count_nomina}')
    print(f'Registros de seguridad social importados: {count_seg}')


if __name__ == '__main__':
    migrate_nomina()
