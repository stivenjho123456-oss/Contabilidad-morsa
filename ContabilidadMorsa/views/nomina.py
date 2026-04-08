import os
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk

from database import (
    AppValidationError,
    delete_nomina_novedad,
    delete_nomina_resumen,
    delete_nomina_seg_social,
    get_nomina_novedades,
    get_nomina_periodos,
    get_nomina_resumen,
    get_nomina_seg_social,
    get_nomina_stats,
    save_nomina_novedad,
    save_nomina_resumen,
    save_nomina_seg_social,
    sync_nomina_to_egresos,
)
from migrate_nomina import EXCEL_NOMINA_PATH, migrate_nomina
from ui_helpers import fit_tree_rows, fix_scrollframe, setup_toplevel, setup_treeview_style

EMP_COLS = (
    'empleado', 'cedula', 'valor_dia', 'q1_dias', 'q1_neto', 'q2_dias', 'q2_neto',
    'total_deduccion', 'total_mes',
)
EMP_LABELS = (
    'Empleado', 'Cedula', 'Valor Dia', 'Dias Q1', 'Neto Q1', 'Dias Q2', 'Neto Q2',
    'Deducciones', 'Total Mes',
)
EMP_WIDTHS = (220, 110, 110, 80, 110, 80, 110, 110, 120)

NOV_COLS = ('fecha', 'empleado', 'quincena', 'naturaleza', 'tipo_novedad', 'valor', 'observaciones')
NOV_LABELS = ('Fecha', 'Empleado', 'Quincena', 'Naturaleza', 'Tipo', 'Valor', 'Observaciones')
NOV_WIDTHS = (95, 200, 90, 100, 120, 110, 200)

SEG_COLS = ('grupo', 'concepto', 'valor', 'observaciones')
SEG_LABELS = ('Grupo', 'Concepto', 'Valor', 'Observaciones')
SEG_WIDTHS = (160, 180, 120, 220)

NATURALEZAS = ['DEVENGADO', 'DEDUCCION']
TIPOS_NOVEDAD = ['BONIFICACION', 'AUXILIO', 'INCAPACIDAD', 'PRESTAMO', 'DESCUENTO', 'OTRO']
QUINCENAS = ['Q1', 'Q2', 'MES']
GRUPOS_SEG = ['SALUD', 'PENSION', 'ARL', 'CAJA COMPENSACION', 'OTRO']


def fmt(value):
    try:
        return f'$ {float(value):,.0f}'.replace(',', '.')
    except Exception:
        return '$ 0'


def _make_style(name, bg_heading):
    palette = {
        'Emp': ('#dbeafe', '#1e3a5f', '#bfdbfe'),
        'Nov': ('#e0f2fe', '#0369a1', '#bae6fd'),
        'Seg': ('#ede9fe', '#6d28d9', '#ddd6fe'),
    }
    heading_bg, heading_fg, select_bg = palette.get(name, ('#e2e8f0', '#334155', '#dbeafe'))
    setup_treeview_style(name, heading_bg=heading_bg, heading_fg=heading_fg, select_bg=select_bg)


class NominaView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color='transparent')
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=2)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)

        # Register styles after theme_use('clam') has been called in main()
        _make_style('Emp', '#1e3a5f')
        _make_style('Nov', '#2563eb')
        _make_style('Seg', '#7c3aed')

        self._resumen_data = []
        self._novedades = []
        self._seg_data = []

        self._build_toolbar()
        self._build_cards()
        self._build_tables()
        self.refresh()

    # ── Toolbar ──────────────────────────────────────────────────────────────

    def _build_toolbar(self):
        bar = ctk.CTkFrame(self, fg_color='white', corner_radius=12)
        bar.grid(row=0, column=0, sticky='ew', pady=(0, 12))

        ctk.CTkLabel(bar, text='Nomina',
                     font=ctk.CTkFont(size=22, weight='bold'),
                     text_color='#1e3a5f').pack(side='left', padx=16, pady=12)

        ctrl = ctk.CTkFrame(bar, fg_color='transparent')
        ctrl.pack(side='right', padx=12)

        self._search_var = tk.StringVar()
        self._search_var.trace_add('write', lambda *_: self.refresh())
        ctk.CTkEntry(ctrl, textvariable=self._search_var,
                     placeholder_text='Buscar empleado o cedula...',
                     width=200, height=32).pack(side='left', padx=(0, 8))

        periodos = get_nomina_periodos() or ['Sin periodos']
        self._periodo_var = tk.StringVar(value=periodos[0])
        self._periodo_cb = ctk.CTkComboBox(ctrl, values=periodos,
                                            variable=self._periodo_var,
                                            width=160, height=32,
                                            command=lambda *_: self.refresh())
        self._periodo_cb.pack(side='left', padx=(0, 8))

        def sep():
            ctk.CTkFrame(ctrl, width=2, height=28, fg_color='#e2e8f0').pack(side='left', padx=6)

        self._btn_group(ctrl, [
            ('+ Empleado',    '#1e3a5f', '#2a5298', self._new_empleado),
            ('Editar emp.',   '#0f766e', '#115e59', self._edit_empleado),
            ('Eliminar emp.', '#dc2626', '#b91c1c', self._delete_empleado),
        ])
        sep()
        self._btn_group(ctrl, [
            ('+ Seg. Social',  '#7c3aed', '#6d28d9', self._new_seg),
            ('Editar seg.',    '#7c3aed', '#6d28d9', self._edit_seg),
            ('Eliminar seg.',  '#dc2626', '#b91c1c', self._delete_seg),
        ])
        sep()
        self._btn_group(ctrl, [
            ('+ Novedad',       '#2563eb', '#1d4ed8', self._new_novedad),
            ('Editar novedad',  '#2563eb', '#1d4ed8', self._edit_novedad),
            ('Eliminar nov.',   '#dc2626', '#b91c1c', self._delete_novedad),
        ])
        sep()
        self._btn_group(ctrl, [
            ('Importar Excel',   '#64748b', '#475569', self._import_nomina),
            ('Sincronizar egr.', '#334155', '#1e293b', self._sync_egresos),
            ('Exportar',         '#16a34a', '#15803d', self._export_nomina),
        ])

    def _btn_group(self, parent, buttons):
        for text, fg, hover, cmd in buttons:
            ctk.CTkButton(parent, text=text, width=110, height=32,
                          fg_color=fg, hover_color=hover,
                          command=cmd).pack(side='left', padx=(0, 4))

    # ── Cards ─────────────────────────────────────────────────────────────────

    def _build_cards(self):
        row = ctk.CTkFrame(self, fg_color='transparent')
        row.grid(row=1, column=0, sticky='ew', pady=(0, 12))
        row.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self._c_empleados = self._card(row, 0, 'Empleados',        '#1e3a5f')
        self._c_nomina    = self._card(row, 1, 'Total Nomina',     '#16a34a')
        self._c_nov_dev   = self._card(row, 2, 'Novedades +',      '#2563eb')
        self._c_nov_ded   = self._card(row, 3, 'Novedades -',      '#dc2626')
        self._c_seg       = self._card(row, 4, 'Seg. Social',      '#7c3aed')
        self._c_integrada = self._card(row, 5, 'Nomina Integrada', '#0f766e')

    def _card(self, parent, col, title, color):
        f = ctk.CTkFrame(parent, fg_color='white', corner_radius=12)
        f.grid(row=0, column=col, sticky='ew', padx=(0 if col == 0 else 6, 0))
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=12),
                     text_color='#64748b').pack(anchor='w', padx=14, pady=(12, 2))
        lbl = ctk.CTkLabel(f, text='0', font=ctk.CTkFont(size=19, weight='bold'),
                           text_color=color)
        lbl.pack(anchor='w', padx=14, pady=(0, 12))
        return lbl

    # ── Tables ────────────────────────────────────────────────────────────────

    def _build_tables(self):
        self._emp_tree = self._build_section(
            grid_row=2, title='Empleados en Nomina',
            style='Emp', columns=EMP_COLS, labels=EMP_LABELS, widths=EMP_WIDTHS)
        self._emp_tree.bind('<Double-1>', lambda _: self._edit_empleado())

        self._seg_tree = self._build_section(
            grid_row=3, title='Seguridad Social',
            style='Seg', columns=SEG_COLS, labels=SEG_LABELS, widths=SEG_WIDTHS)
        self._seg_tree.bind('<Double-1>', lambda _: self._edit_seg())

        self._nov_tree = self._build_section(
            grid_row=4, title='Novedades Manuales',
            style='Nov', columns=NOV_COLS, labels=NOV_LABELS, widths=NOV_WIDTHS)
        self._nov_tree.bind('<Double-1>', lambda _: self._edit_novedad())

    def _build_section(self, grid_row, title, style, columns, labels, widths):
        frame = ctk.CTkFrame(self, fg_color='white', corner_radius=12)
        frame.grid(row=grid_row, column=0, sticky='nsew',
                   pady=(0, 0 if grid_row == 4 else 10))
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(frame, text=title,
                     font=ctk.CTkFont(size=14, weight='bold'),
                     text_color='#1e3a5f').grid(row=0, column=0, sticky='w',
                                                padx=16, pady=(12, 4))

        tree = ttk.Treeview(frame, style=f'{style}.Treeview',
                             columns=columns, show='headings', selectmode='browse')
        for col, lbl, w in zip(columns, labels, widths):
            tree.heading(col, text=lbl)
            anchor = 'w' if col in ('empleado', 'observaciones', 'concepto') else 'center'
            tree.column(col, width=w, minwidth=60, anchor=anchor)

        vsb = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=1, column=0, sticky='ew', padx=8, pady=4)
        vsb.grid(row=1, column=1, sticky='ns', pady=4)
        hsb.grid(row=2, column=0, sticky='ew', padx=8)
        return tree

    # ── Data ─────────────────────────────────────────────────────────────────

    def _periodo(self):
        v = self._periodo_var.get()
        return None if v in ('', 'Sin periodos', 'Todos') else v

    def refresh(self):
        periodos = get_nomina_periodos()
        values = periodos if periodos else ['Sin periodos']
        self._periodo_cb.configure(values=values)
        if self._periodo_var.get() not in values:
            self._periodo_var.set(values[0])

        periodo = self._periodo()
        search = self._search_var.get()

        self._resumen_data = get_nomina_resumen(periodo=periodo, search=search)
        self._seg_data = get_nomina_seg_social(periodo=periodo)
        self._novedades = get_nomina_novedades(periodo=periodo, search=search)
        stats = get_nomina_stats(periodo=periodo)

        self._c_empleados.configure(text=str(stats['empleados']))
        self._c_nomina.configure(text=fmt(stats['total_nomina']))
        self._c_nov_dev.configure(text=fmt(stats['total_novedades_devengado']))
        self._c_nov_ded.configure(text=fmt(stats['total_novedades_deduccion']))
        self._c_seg.configure(text=fmt(stats['total_seg_social']))
        self._c_integrada.configure(text=fmt(stats['total_nomina_integrada']))

        self._fill(self._emp_tree,
                   [(r['empleado'], r['cedula'], fmt(r['valor_dia']),
                     r['q1_dias'], fmt(r['q1_neto']),
                     r['q2_dias'], fmt(r['q2_neto']),
                     fmt(r['total_deduccion']), fmt(r['total_mes']))
                    for r in self._resumen_data],
                   ids=[str(r['id']) for r in self._resumen_data])

        self._fill(self._seg_tree,
                   [(r['grupo'] or '', r['concepto'], fmt(r['valor']),
                     r['observaciones'] or '')
                    for r in self._seg_data],
                   ids=[str(r['id']) for r in self._seg_data])

        self._fill(self._nov_tree,
                   [(r['fecha'], r['empleado'], r['quincena'], r['naturaleza'],
                     r['tipo_novedad'], fmt(r['valor']), r['observaciones'] or '')
                    for r in self._novedades],
                   ids=[str(r['id']) for r in self._novedades])

    def _fill(self, tree, rows, ids=None):
        tree.delete(*tree.get_children())
        for i, values in enumerate(rows):
            tag = 'even' if i % 2 == 0 else 'odd'
            tree.insert('', 'end', iid=ids[i] if ids else None,
                        tags=(tag,), values=values)
        fit_tree_rows(tree, len(rows), max_rows=12)
        tree.tag_configure('even', background='white', foreground='#1f2937')
        tree.tag_configure('odd', background='#f8fafc', foreground='#1f2937')

    def _sel(self, tree):
        s = tree.selection()
        return int(s[0]) if s else None

    # ── Empleados CRUD ────────────────────────────────────────────────────────

    def _new_empleado(self):
        EmpleadoForm(self, None, self._periodo() or '', self._after_emp_saved)

    def _edit_empleado(self):
        eid = self._sel(self._emp_tree)
        if eid is None:
            messagebox.showinfo('Aviso', 'Selecciona un empleado para editar.')
            return
        data = next((r for r in self._resumen_data if r['id'] == eid), None)
        if data:
            EmpleadoForm(self, data, data['periodo'], self._after_emp_saved)

    def _delete_empleado(self):
        eid = self._sel(self._emp_tree)
        if eid is None:
            messagebox.showinfo('Aviso', 'Selecciona un empleado para eliminar.')
            return
        emp = next((r for r in self._resumen_data if r['id'] == eid), None)
        nombre = emp['empleado'] if emp else ''
        if messagebox.askyesno('Confirmar', f'¿Eliminar a {nombre} de la nomina?'):
            delete_nomina_resumen(eid)
            sync_nomina_to_egresos(self._periodo())
            self.refresh()

    def _after_emp_saved(self, periodo):
        sync_nomina_to_egresos(periodo or None)
        self.refresh()

    # ── Seguridad Social CRUD ─────────────────────────────────────────────────

    def _new_seg(self):
        periodo = self._periodo()
        if not periodo:
            messagebox.showinfo('Aviso', 'Selecciona un periodo primero.')
            return
        SegSocialForm(self, None, periodo, self.refresh)

    def _edit_seg(self):
        sid = self._sel(self._seg_tree)
        if sid is None:
            messagebox.showinfo('Aviso', 'Selecciona un item de seg. social para editar.')
            return
        data = next((r for r in self._seg_data if r['id'] == sid), None)
        if data:
            SegSocialForm(self, data, data['periodo'], self.refresh)

    def _delete_seg(self):
        sid = self._sel(self._seg_tree)
        if sid is None:
            messagebox.showinfo('Aviso', 'Selecciona un item de seg. social para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar este registro de seguridad social?'):
            delete_nomina_seg_social(sid)
            self.refresh()

    # ── Novedades CRUD ────────────────────────────────────────────────────────

    def _new_novedad(self):
        empleados = [{'empleado': r['empleado'], 'cedula': r['cedula']}
                     for r in self._resumen_data]
        NovedadForm(self, None, self._periodo() or '', empleados, self._after_nov_saved)

    def _edit_novedad(self):
        nid = self._sel(self._nov_tree)
        if nid is None:
            messagebox.showinfo('Aviso', 'Selecciona una novedad para editar.')
            return
        data = next((n for n in self._novedades if n['id'] == nid), None)
        if data:
            empleados = [{'empleado': r['empleado'], 'cedula': r['cedula']}
                         for r in self._resumen_data]
            NovedadForm(self, data, data['periodo'], empleados, self._after_nov_saved)

    def _delete_novedad(self):
        nid = self._sel(self._nov_tree)
        if nid is None:
            messagebox.showinfo('Aviso', 'Selecciona una novedad para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar esta novedad?'):
            delete_nomina_novedad(nid)
            sync_nomina_to_egresos(self._periodo())
            self.refresh()

    def _after_nov_saved(self, periodo):
        sync_nomina_to_egresos(periodo or None)
        self.refresh()

    # ── Import / Sync / Export ────────────────────────────────────────────────

    def _import_nomina(self):
        path = filedialog.askopenfilename(
            title='Selecciona el archivo de nomina',
            filetypes=[('Excel', '*.xlsx *.xls')],
        )
        if not path:
            return
        try:
            migrate_nomina(path=path)
            total = sync_nomina_to_egresos()
        except Exception as exc:
            messagebox.showerror('Error', f'No se pudo importar la nomina.\n\n{exc}')
            return
        self.refresh()
        messagebox.showinfo('Exito', f'Nomina importada correctamente.\n{total} egresos generados.')

    def _sync_egresos(self):
        try:
            total = sync_nomina_to_egresos(periodo=self._periodo())
        except Exception as exc:
            messagebox.showerror('Error', f'No se pudieron sincronizar los egresos.\n\n{exc}')
            return
        messagebox.showinfo('Exito', f'Se generaron {total} egresos automaticos de nomina.')

    def _export_nomina(self):
        periodo = self._periodo()
        if not periodo:
            messagebox.showinfo('Aviso', 'Selecciona un periodo especifico para exportar.')
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialfile=f'Nomina_{periodo.replace(" ", "_")}.xlsx',
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messagebox.showerror('Error', 'Instala openpyxl: pip install openpyxl')
            return

        resumen = get_nomina_resumen(periodo=periodo)
        novedades = get_nomina_novedades(periodo=periodo)
        seg = get_nomina_seg_social(periodo=periodo)
        stats = get_nomina_stats(periodo=periodo)

        wb = openpyxl.Workbook()
        hdr_font = Font(color='FFFFFF', bold=True)
        center = Alignment(horizontal='center')

        def fill(color):
            return PatternFill('solid', fgColor=color)

        ws = wb.active
        ws.title = 'Resumen'
        ws['A1'] = f'NOMINA {periodo}'
        ws['A1'].font = Font(bold=True, size=14)
        for i, (lbl, val) in enumerate([
            ('Empleados', stats['empleados']),
            ('Total nomina', stats['total_nomina']),
            ('Novedades devengado', stats['total_novedades_devengado']),
            ('Novedades deduccion', stats['total_novedades_deduccion']),
            ('Seguridad social', stats['total_seg_social']),
            ('Nomina integrada', stats['total_nomina_integrada']),
        ], start=3):
            ws.cell(i, 1, lbl)
            ws.cell(i, 2, val).number_format = '$ #,##0'
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18

        ws_emp = wb.create_sheet('Empleados')
        for col, lbl in enumerate(EMP_LABELS, 1):
            c = ws_emp.cell(1, col, lbl)
            c.fill = fill('1e3a5f'); c.font = hdr_font; c.alignment = center
        for r, row in enumerate(resumen, 2):
            ws_emp.cell(r, 1, row['empleado'])
            ws_emp.cell(r, 2, row['cedula'])
            ws_emp.cell(r, 3, row['valor_dia']).number_format = '$ #,##0'
            ws_emp.cell(r, 4, row['q1_dias'])
            ws_emp.cell(r, 5, row['q1_neto']).number_format = '$ #,##0'
            ws_emp.cell(r, 6, row['q2_dias'])
            ws_emp.cell(r, 7, row['q2_neto']).number_format = '$ #,##0'
            ws_emp.cell(r, 8, row['total_deduccion']).number_format = '$ #,##0'
            ws_emp.cell(r, 9, row['total_mes']).number_format = '$ #,##0'

        ws_seg = wb.create_sheet('SegSocial')
        for col, lbl in enumerate(SEG_LABELS, 1):
            c = ws_seg.cell(1, col, lbl)
            c.fill = fill('7c3aed'); c.font = hdr_font; c.alignment = center
        for r, row in enumerate(seg, 2):
            ws_seg.cell(r, 1, row['grupo'] or '')
            ws_seg.cell(r, 2, row['concepto'])
            ws_seg.cell(r, 3, row['valor']).number_format = '$ #,##0'
            ws_seg.cell(r, 4, row['observaciones'] or '')

        ws_nov = wb.create_sheet('Novedades')
        for col, lbl in enumerate(NOV_LABELS, 1):
            c = ws_nov.cell(1, col, lbl)
            c.fill = fill('2563eb'); c.font = hdr_font; c.alignment = center
        for r, row in enumerate(novedades, 2):
            ws_nov.cell(r, 1, row['fecha'])
            ws_nov.cell(r, 2, row['empleado'])
            ws_nov.cell(r, 3, row['quincena'])
            ws_nov.cell(r, 4, row['naturaleza'])
            ws_nov.cell(r, 5, row['tipo_novedad'])
            ws_nov.cell(r, 6, row['valor']).number_format = '$ #,##0'
            ws_nov.cell(r, 7, row['observaciones'] or '')

        try:
            wb.save(path)
        except PermissionError:
            messagebox.showerror('Error', 'Cierra el Excel si lo tienes abierto e intenta de nuevo.')
            return
        messagebox.showinfo('Exito', f'Reporte exportado:\n{path}')


# ── Formulario Empleado ───────────────────────────────────────────────────────

class EmpleadoForm(ctk.CTkToplevel):
    def __init__(self, parent, data, periodo_default, on_save):
        super().__init__(parent)
        self.on_save = on_save
        self.editing_id = data['id'] if data else None
        self.title('Nuevo Empleado en Nomina' if not data else 'Editar Empleado')
        self.geometry('560x760')
        self.resizable(False, True)
        self.transient(parent.winfo_toplevel())

        # Botones al fondo primero — siempre visibles
        btn_f = ctk.CTkFrame(self, fg_color='white')
        btn_f.pack(side='bottom', fill='x', padx=24, pady=(0, 20))
        ctk.CTkButton(btn_f, text='Cancelar', width=100, height=36,
                      fg_color='#cbd5e1', text_color='#0f172a', hover_color='#94a3b8',
                      command=self.destroy).pack(side='right', padx=(8, 0))
        ctk.CTkButton(btn_f, text='Guardar', width=120, height=36,
                      fg_color='#1e3a5f', hover_color='#2a5298',
                      command=self._save).pack(side='right')

        self._build(data or {}, periodo_default)
        self.after(100, lambda: (self.grab_set(), self.focus_force()))

    def _build(self, d, periodo_default):
        setup_toplevel(self)
        main = ctk.CTkScrollableFrame(self, fg_color='white')
        main.pack(fill='both', expand=True, padx=2, pady=2)
        fix_scrollframe(main)

        ctk.CTkLabel(main, text='Datos del Empleado',
                     font=ctk.CTkFont(size=16, weight='bold'),
                     text_color='#1e3a5f').pack(anchor='w', padx=24, pady=(16, 4))

        def lbl(text):
            ctk.CTkLabel(main, text=text, font=ctk.CTkFont(size=12),
                         text_color='#475569').pack(anchor='w', padx=24, pady=(8, 0))

        def entry(default=''):
            e = ctk.CTkEntry(main, height=34)
            e.pack(fill='x', padx=24, pady=(2, 0))
            if default not in ('', None):
                e.insert(0, str(default))
            return e

        def section(text):
            f = ctk.CTkFrame(main, fg_color='#f1f5f9', corner_radius=8)
            f.pack(fill='x', padx=24, pady=(14, 0))
            ctk.CTkLabel(f, text=text,
                         font=ctk.CTkFont(size=13, weight='bold'),
                         text_color='#334155').pack(anchor='w', padx=12, pady=6)

        lbl('Periodo  (ej: MARZO 2025)')
        self._periodo = entry(d.get('periodo', periodo_default))

        lbl('Nombre completo del empleado *')
        self._empleado = entry(d.get('empleado', ''))

        lbl('Cedula')
        self._cedula = entry(d.get('cedula', ''))

        lbl('Valor dia ($)')
        self._valor_dia = entry(int(d['valor_dia']) if d.get('valor_dia') else '')

        # ── Primera quincena ──────────────────────────────────────────────
        section('Primera Quincena (Q1)')

        lbl('Dias trabajados Q1')
        self._q1_dias = entry(d.get('q1_dias', ''))
        self._q1_dias.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Devengado Q1 ($) — salario bruto de la quincena')
        self._q1_dev = entry(int(d['q1_devengado']) if d.get('q1_devengado') else '')
        self._q1_dev.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Auxilio de transporte Q1 ($)')
        self._q1_aux = entry(int(d['q1_aux_transporte']) if d.get('q1_aux_transporte') else '0')
        self._q1_aux.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Descuento salud Q1 ($)')
        self._q1_salud = entry(int(d['q1_salud']) if d.get('q1_salud') else '0')
        self._q1_salud.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Descuento pension Q1 ($)')
        self._q1_pension = entry(int(d['q1_pension']) if d.get('q1_pension') else '0')
        self._q1_pension.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Neto pagado Q1 ($) — se calcula automaticamente')
        self._q1_neto = entry(int(d['q1_neto']) if d.get('q1_neto') else '0')

        # ── Segunda quincena ──────────────────────────────────────────────
        section('Segunda Quincena (Q2)')

        lbl('Dias trabajados Q2')
        self._q2_dias = entry(d.get('q2_dias', ''))
        self._q2_dias.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Devengado Q2 ($)')
        self._q2_dev = entry(int(d['q2_devengado']) if d.get('q2_devengado') else '')
        self._q2_dev.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Auxilio de transporte Q2 ($)')
        self._q2_aux = entry(int(d['q2_aux_transporte']) if d.get('q2_aux_transporte') else '0')
        self._q2_aux.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Descuento salud Q2 ($)')
        self._q2_salud = entry(int(d['q2_salud']) if d.get('q2_salud') else '0')
        self._q2_salud.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Descuento pension Q2 ($)')
        self._q2_pension = entry(int(d['q2_pension']) if d.get('q2_pension') else '0')
        self._q2_pension.bind('<FocusOut>', lambda _: self._recalcular())

        lbl('Neto pagado Q2 ($) — se calcula automaticamente')
        self._q2_neto = entry(int(d['q2_neto']) if d.get('q2_neto') else '0')

        # ── Totales ───────────────────────────────────────────────────────
        section('Totales del Mes')

        lbl('Total deducciones ($) — se calcula automaticamente')
        self._total_ded = entry(int(d['total_deduccion']) if d.get('total_deduccion') else '0')

        lbl('Total incapacidades ($)')
        self._total_inc = entry(int(d['total_incapacidad']) if d.get('total_incapacidad') else '0')

        lbl('Total descuentos adicionales ($)')
        self._total_desc = entry(int(d['total_descuento']) if d.get('total_descuento') else '0')

        lbl('Total mes ($) — se calcula automaticamente')
        self._total_mes = entry(int(d['total_mes']) if d.get('total_mes') else '0')

    def _n(self, widget):
        raw = widget.get().strip().replace('.', '').replace(',', '').replace('$', '').replace(' ', '')
        try:
            return float(raw) if raw else 0.0
        except ValueError:
            return 0.0

    def _set(self, widget, value):
        widget.delete(0, 'end')
        widget.insert(0, str(int(value)))

    def _recalcular(self):
        q1_neto = (self._n(self._q1_dev) + self._n(self._q1_aux)
                   - self._n(self._q1_salud) - self._n(self._q1_pension))
        q2_neto = (self._n(self._q2_dev) + self._n(self._q2_aux)
                   - self._n(self._q2_salud) - self._n(self._q2_pension))
        total_ded = (self._n(self._q1_salud) + self._n(self._q1_pension)
                     + self._n(self._q2_salud) + self._n(self._q2_pension))
        total_mes = q1_neto + q2_neto

        self._set(self._q1_neto, max(q1_neto, 0))
        self._set(self._q2_neto, max(q2_neto, 0))
        self._set(self._total_ded, total_ded)
        self._set(self._total_mes, max(total_mes, 0))

    def _save(self):
        periodo = self._periodo.get().strip().upper()
        empleado = self._empleado.get().strip()
        if not periodo or not empleado:
            messagebox.showerror('Error', 'Periodo y nombre del empleado son obligatorios.')
            return
        self._recalcular()
        data = {
            'periodo':           periodo,
            'empleado':          empleado,
            'cedula':            self._cedula.get().strip(),
            'valor_dia':         self._n(self._valor_dia),
            'q1_dias':           self._n(self._q1_dias),
            'q1_devengado':      self._n(self._q1_dev),
            'q1_aux_transporte': self._n(self._q1_aux),
            'q1_salud':          self._n(self._q1_salud),
            'q1_pension':        self._n(self._q1_pension),
            'q1_neto':           self._n(self._q1_neto),
            'q2_dias':           self._n(self._q2_dias),
            'q2_devengado':      self._n(self._q2_dev),
            'q2_aux_transporte': self._n(self._q2_aux),
            'q2_salud':          self._n(self._q2_salud),
            'q2_pension':        self._n(self._q2_pension),
            'q2_neto':           self._n(self._q2_neto),
            'total_deduccion':   self._n(self._total_ded),
            'total_incapacidad': self._n(self._total_inc),
            'total_descuento':   self._n(self._total_desc),
            'total_mes':         self._n(self._total_mes),
            'origen_archivo':    'MANUAL',
        }
        save_nomina_resumen(data, self.editing_id)
        self.destroy()
        self.on_save(periodo)


# ── Formulario Seguridad Social ───────────────────────────────────────────────

class SegSocialForm(ctk.CTkToplevel):
    def __init__(self, parent, data, periodo, on_save):
        super().__init__(parent)
        self.on_save = on_save
        self.editing_id = data['id'] if data else None
        self.periodo = periodo
        self.title('Nuevo Aporte Seg. Social' if not data else 'Editar Seg. Social')
        self.geometry('460x440')
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())

        btn_f = ctk.CTkFrame(self, fg_color='white')
        btn_f.pack(side='bottom', fill='x', padx=24, pady=(0, 20))
        ctk.CTkButton(btn_f, text='Cancelar', width=100, height=36,
                      fg_color='#cbd5e1', text_color='#0f172a', hover_color='#94a3b8',
                      command=self.destroy).pack(side='right', padx=(8, 0))
        ctk.CTkButton(btn_f, text='Guardar', width=120, height=36,
                      fg_color='#7c3aed', hover_color='#6d28d9',
                      command=self._save).pack(side='right')

        self._build(data or {})
        self.after(100, lambda: (self.grab_set(), self.focus_force()))

    def _build(self, d):
        setup_toplevel(self)
        main = ctk.CTkFrame(self, fg_color='white')
        main.pack(fill='both', expand=True, padx=24, pady=20)

        ctk.CTkLabel(main, text='Seguridad Social',
                     font=ctk.CTkFont(size=16, weight='bold'),
                     text_color='#7c3aed').pack(anchor='w', pady=(0, 12))

        def lbl(text):
            ctk.CTkLabel(main, text=text, font=ctk.CTkFont(size=12),
                         text_color='#475569').pack(anchor='w', pady=(8, 0))

        lbl('Periodo')
        self._periodo_e = ctk.CTkEntry(main, height=34)
        self._periodo_e.pack(fill='x')
        self._periodo_e.insert(0, d.get('periodo', self.periodo))

        lbl('Grupo')
        self._grupo_var = tk.StringVar(value=d.get('grupo', 'SALUD') or 'SALUD')
        ctk.CTkComboBox(main, values=GRUPOS_SEG, variable=self._grupo_var, height=34).pack(fill='x')

        lbl('Concepto')
        self._concepto = ctk.CTkEntry(main, height=34)
        self._concepto.pack(fill='x')
        self._concepto.insert(0, d.get('concepto', ''))

        lbl('Valor ($)')
        self._valor = ctk.CTkEntry(main, height=34)
        self._valor.pack(fill='x')
        if d.get('valor'):
            self._valor.insert(0, str(int(d['valor'])))

        lbl('Observaciones')
        self._obs = ctk.CTkEntry(main, height=34)
        self._obs.pack(fill='x')
        self._obs.insert(0, d.get('observaciones', '') or '')

    def _n(self, widget):
        raw = widget.get().strip().replace('.', '').replace(',', '').replace('$', '').replace(' ', '')
        try:
            return float(raw) if raw else 0.0
        except ValueError:
            return 0.0

    def _save(self):
        periodo = self._periodo_e.get().strip().upper()
        concepto = self._concepto.get().strip()
        if not periodo or not concepto:
            messagebox.showerror('Error', 'Periodo y concepto son obligatorios.')
            return
        data = {
            'periodo':        periodo,
            'grupo':          self._grupo_var.get().strip().upper(),
            'concepto':       concepto.upper(),
            'valor':          self._n(self._valor),
            'observaciones':  self._obs.get().strip(),
            'origen_archivo': 'MANUAL',
        }
        save_nomina_seg_social(data, self.editing_id)
        self.destroy()
        self.on_save()


# ── Formulario Novedad ────────────────────────────────────────────────────────

class NovedadForm(ctk.CTkToplevel):
    def __init__(self, parent, data, periodo_default, empleados, on_save):
        super().__init__(parent)
        self.on_save = on_save
        self.editing_id = data['id'] if data else None
        self.empleados = empleados
        self.periodo_default = periodo_default or ''

        self.title('Nueva Novedad' if not data else 'Editar Novedad')
        self.geometry('520x600')
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())

        btn_f = ctk.CTkFrame(self, fg_color='white')
        btn_f.pack(side='bottom', fill='x', padx=24, pady=(0, 20))
        ctk.CTkButton(btn_f, text='Cancelar', width=100, height=36,
                      fg_color='#cbd5e1', text_color='#0f172a', hover_color='#94a3b8',
                      command=self.destroy).pack(side='right', padx=(8, 0))
        ctk.CTkButton(btn_f, text='Guardar', width=120, height=36,
                      fg_color='#2563eb', hover_color='#1d4ed8',
                      command=self._save).pack(side='right')

        self._build(data or {})
        self.after(100, lambda: (self.grab_set(), self.focus_force()))

    def _build(self, d):
        setup_toplevel(self)
        main = ctk.CTkScrollableFrame(self, fg_color='white')
        main.pack(fill='both', expand=True, padx=2, pady=2)
        fix_scrollframe(main)

        ctk.CTkLabel(main, text='Novedad de Nomina',
                     font=ctk.CTkFont(size=16, weight='bold'),
                     text_color='#2563eb').pack(anchor='w', padx=24, pady=(16, 4))

        def lbl(text):
            ctk.CTkLabel(main, text=text, font=ctk.CTkFont(size=12),
                         text_color='#475569').pack(anchor='w', padx=24, pady=(8, 0))

        def entry(default=''):
            e = ctk.CTkEntry(main, height=34)
            e.pack(fill='x', padx=24, pady=(2, 0))
            if default:
                e.insert(0, str(default))
            return e

        lbl('Periodo')
        self._periodo = entry(d.get('periodo', self.periodo_default))

        lbl('Fecha')
        self._fecha = entry(d.get('fecha', date.today().isoformat()))

        lbl('Empleado')
        nombres = [e['empleado'] for e in self.empleados]
        self._empleado_var = tk.StringVar(value=d.get('empleado', ''))
        self._empleado_cb = ctk.CTkComboBox(main, values=nombres,
                                             variable=self._empleado_var, height=34)
        self._empleado_cb.pack(fill='x', padx=24, pady=(2, 0))
        self._empleado_var.trace_add('write', self._on_emp_change)

        lbl('Cedula')
        self._cedula = entry(d.get('cedula', ''))

        lbl('Quincena')
        self._quincena_var = tk.StringVar(value=d.get('quincena', 'MES') or 'MES')
        ctk.CTkComboBox(main, values=QUINCENAS, variable=self._quincena_var, height=34).pack(
            fill='x', padx=24, pady=(2, 0))

        lbl('Naturaleza')
        self._naturaleza_var = tk.StringVar(value=d.get('naturaleza', 'DEVENGADO') or 'DEVENGADO')
        ctk.CTkComboBox(main, values=NATURALEZAS, variable=self._naturaleza_var, height=34).pack(
            fill='x', padx=24, pady=(2, 0))

        lbl('Tipo de novedad')
        self._tipo_var = tk.StringVar(value=d.get('tipo_novedad', 'BONIFICACION') or 'BONIFICACION')
        ctk.CTkComboBox(main, values=TIPOS_NOVEDAD, variable=self._tipo_var, height=34).pack(
            fill='x', padx=24, pady=(2, 0))

        lbl('Valor ($)')
        self._valor = entry(str(int(d['valor'])) if d.get('valor') else '')

        lbl('Observaciones')
        self._obs = entry(d.get('observaciones', ''))

    def _on_emp_change(self, *_):
        nombre = self._empleado_var.get().strip().upper()
        found = next((e for e in self.empleados
                      if e['empleado'].strip().upper() == nombre), None)
        if found:
            self._cedula.delete(0, 'end')
            self._cedula.insert(0, found.get('cedula', '') or '')

    def _n(self, widget):
        raw = widget.get().strip().replace('.', '').replace(',', '').replace('$', '').replace(' ', '')
        try:
            return float(raw) if raw else 0.0
        except ValueError:
            return 0.0

    def _save(self):
        periodo = self._periodo.get().strip().upper()
        fecha = self._fecha.get().strip()
        empleado = self._empleado_var.get().strip()
        if not periodo or not fecha or not empleado:
            messagebox.showerror('Error', 'Periodo, fecha y empleado son obligatorios.')
            return
        data = {
            'periodo':        periodo,
            'fecha':          fecha,
            'empleado':       empleado,
            'cedula':         self._cedula.get().strip(),
            'quincena':       self._quincena_var.get().strip(),
            'naturaleza':     self._naturaleza_var.get().strip().upper(),
            'tipo_novedad':   self._tipo_var.get().strip().upper(),
            'valor':          self._n(self._valor),
            'observaciones':  self._obs.get().strip(),
            'origen_archivo': 'MANUAL',
        }
        try:
            save_nomina_novedad(data, self.editing_id)
        except AppValidationError as exc:
            messagebox.showerror('Error', str(exc))
            return
        self.destroy()
        self.on_save(periodo)
