from collections import defaultdict
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from ui_helpers import fit_tree_rows, fix_scrollframe, setup_treeview_style

from database import (
    get_cierre_mensual,
    get_egresos,
    get_ingresos,
    get_nomina_novedades,
    get_nomina_resumen,
    get_nomina_seg_social,
    get_nomina_stats,
    get_proveedores,
    period_from_month_year,
)

MONTHS = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}
MONTH_NAMES = [MONTHS[i] for i in range(1, 13)]


def fmt(v):
    try:
        return f'$ {float(v):,.0f}'.replace(',', '.')
    except Exception:
        return '$ 0'


class ReportesView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color='transparent')
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        now = datetime.now()
        self._mes = now.month
        self._ano = now.year

        self._build_toolbar()
        self._build_content()
        self.refresh()

    def _build_toolbar(self):
        bar = ctk.CTkFrame(self, fg_color='white', corner_radius=12)
        bar.grid(row=0, column=0, sticky='ew', pady=(0, 12))

        ctk.CTkLabel(bar, text='Reportes y Cierre', font=ctk.CTkFont(size=22, weight='bold'),
                     text_color='#1e3a5f').pack(side='left', padx=16, pady=12)

        ctrl = ctk.CTkFrame(bar, fg_color='transparent')
        ctrl.pack(side='right', padx=12)

        self._mes_cb = ctk.CTkComboBox(ctrl, values=MONTH_NAMES, width=130, height=32,
                                       command=self._on_period)
        self._mes_cb.set(MONTHS[self._mes])
        self._mes_cb.pack(side='left', padx=(0, 4))

        self._ano_cb = ctk.CTkComboBox(ctrl, values=[str(y) for y in range(2023, 2030)],
                                       width=84, height=32, command=self._on_period)
        self._ano_cb.set(str(self._ano))
        self._ano_cb.pack(side='left', padx=(0, 12))

        ctk.CTkButton(ctrl, text='Exportar Reporte', width=138, height=32,
                      fg_color='#16a34a', hover_color='#15803d',
                      command=self._export).pack(side='left', padx=(0, 6))
        ctk.CTkButton(ctrl, text='Exportar Cierre', width=138, height=32,
                      fg_color='#1d4ed8', hover_color='#1e40af',
                      command=self._export_cierre).pack(side='left', padx=(0, 6))
        ctk.CTkButton(ctrl, text='Exportar Todo', width=138, height=32,
                      fg_color='#7c3aed', hover_color='#6d28d9',
                      command=self._export_all).pack(side='left')

    def _build_content(self):
        content = ctk.CTkFrame(self, fg_color='transparent')
        content.grid(row=1, column=0, sticky='nsew')
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)
        content.grid_rowconfigure(0, weight=1)

        left = ctk.CTkFrame(content, fg_color='white', corner_radius=12)
        left.grid(row=0, column=0, sticky='nsew', padx=(0, 8))
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(left, text='Cierre Mensual Unificado',
                     font=ctk.CTkFont(size=15, weight='bold'),
                     text_color='#1e3a5f').grid(row=0, column=0, sticky='w', padx=18, pady=(14, 6))
        self._resumen_frame = ctk.CTkScrollableFrame(left, fg_color='white')
        fix_scrollframe(self._resumen_frame)
        self._resumen_frame.grid(row=1, column=0, sticky='nsew', padx=8, pady=(0, 8))

        right = ctk.CTkFrame(content, fg_color='white', corner_radius=12)
        right.grid(row=0, column=1, sticky='nsew')
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(right, text='Detalle por Naturaleza del gasto',
                     font=ctk.CTkFont(size=15, weight='bold'),
                     text_color='#1e3a5f').grid(row=0, column=0, sticky='w', padx=18, pady=(14, 6))

        tv_frame = ctk.CTkFrame(right, fg_color='transparent')
        tv_frame.grid(row=1, column=0, sticky='nsew', padx=8, pady=(0, 8))
        tv_frame.grid_columnconfigure(0, weight=1)

        setup_treeview_style('Rep', heading_bg='#ede9fe', heading_fg='#5b21b6', select_bg='#ddd6fe')

        cols = ('tipo', 'cantidad', 'total', 'porcentaje')
        self._tipo_tree = ttk.Treeview(tv_frame, style='Rep.Treeview',
                                       columns=cols, show='headings', selectmode='none')
        for col, label, width, anchor in (
            ('tipo', 'Naturaleza', 170, 'w'),
            ('cantidad', 'N° Registros', 110, 'center'),
            ('total', 'Total', 140, 'e'),
            ('porcentaje', '%', 70, 'center'),
        ):
            self._tipo_tree.heading(col, text=label)
            self._tipo_tree.column(col, width=width, anchor=anchor)
        vsb = ttk.Scrollbar(tv_frame, orient='vertical', command=self._tipo_tree.yview)
        self._tipo_tree.configure(yscrollcommand=vsb.set)
        self._tipo_tree.grid(row=0, column=0, sticky='ew')
        vsb.grid(row=0, column=1, sticky='ns')

    def _on_period(self, *_):
        self._mes = MONTH_NAMES.index(self._mes_cb.get()) + 1
        self._ano = int(self._ano_cb.get())
        self.refresh()

    def refresh(self):
        cierre = get_cierre_mensual(self._mes, self._ano)
        for w in self._resumen_frame.winfo_children():
            w.destroy()

        items = [
            ('Periodo', cierre['periodo'] or f'{MONTHS[self._mes]} {self._ano}', '#1e3a5f', True),
            ('─' * 32, '', '#cbd5e1', False),
            ('INGRESOS', fmt(cierre['total_ingresos']), '#16a34a', True),
            ('Egresos operativos', fmt(cierre['egresos_operativos']), '#334155', False),
            ('Nomina empleados', fmt(cierre['egresos_nomina']), '#2563eb', False),
            ('Seguridad social', fmt(cierre['egresos_seg_social']), '#7c3aed', False),
            ('Novedades deduccion', fmt(cierre['novedades_deduccion']), '#dc2626', False),
            ('TOTAL EGRESOS', fmt(cierre['total_egresos']), '#dc2626', True),
            ('─' * 32, '', '#cbd5e1', False),
            ('Nomina integrada', fmt(cierre['nomina']['total_nomina_integrada']), '#0f766e', True),
            ('Empleados', str(cierre['nomina']['empleados']), '#1e3a5f', False),
            ('Novedades +', fmt(cierre['nomina']['total_novedades_devengado']), '#2563eb', False),
            ('Novedades -', fmt(cierre['nomina']['total_novedades_deduccion']), '#dc2626', False),
            ('RESULTADO NETO', fmt(cierre['resultado_neto']),
             '#16a34a' if cierre['resultado_neto'] >= 0 else '#dc2626', True),
        ]

        for label, value, color, bold in items:
            row = ctk.CTkFrame(self._resumen_frame, fg_color='transparent')
            row.pack(fill='x', pady=2)
            font = ctk.CTkFont(size=13, weight='bold' if bold else 'normal')
            ctk.CTkLabel(row, text=label, font=font, text_color=color, anchor='w').pack(
                side='left', padx=8
            )
            if value:
                ctk.CTkLabel(row, text=value, font=font, text_color=color, anchor='e').pack(
                    side='right', padx=8
                )

        egresos = get_egresos(mes=self._mes, ano=self._ano)
        total_eg = cierre['total_egresos'] or 1
        self._tipo_tree.delete(*self._tipo_tree.get_children())
        counts = defaultdict(int)
        totals = defaultdict(float)
        for row in egresos:
            if (row.get('source_module') or '').upper() == 'NOMINA' and row['tipo_gasto'] == 'EMPLEADO':
                key = 'NOMINA EMPLEADOS'
            elif (row.get('source_module') or '').upper() == 'NOMINA' and row['tipo_gasto'] == 'SEG SOCIAL':
                key = 'NOMINA SEG SOCIAL'
            else:
                key = row['tipo_gasto'] or 'OTRO'
            counts[key] += 1
            totals[key] += row['valor'] or 0

        for i, (tipo, total) in enumerate(sorted(totals.items(), key=lambda x: x[1], reverse=True)):
            pct = total / total_eg * 100
            tag = 'even' if i % 2 == 0 else 'odd'
            self._tipo_tree.insert('', 'end', tags=(tag,),
                                   values=(tipo, counts[tipo], fmt(total), f'{pct:.1f}%'))
        fit_tree_rows(self._tipo_tree, len(totals), max_rows=12)
        self._tipo_tree.tag_configure('even', background='white', foreground='#1f2937')
        self._tipo_tree.tag_configure('odd', background='#f8f9fa', foreground='#1f2937')

    def _export(self):
        self._export_workbook(cierre_only=False)

    def _export_cierre(self):
        self._export_workbook(cierre_only=True)

    def _export_all(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror('Error', 'Instala openpyxl: pip install openpyxl')
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialfile=f'Contabilidad_Morsa_Completo_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        header_fill = PatternFill('solid', fgColor='1e3a5f')
        green_fill = PatternFill('solid', fgColor='16a34a')
        red_fill = PatternFill('solid', fgColor='dc2626')
        blue_fill = PatternFill('solid', fgColor='1d4ed8')
        purple_fill = PatternFill('solid', fgColor='7c3aed')
        slate_fill = PatternFill('solid', fgColor='334155')
        white_font = Font(color='FFFFFF', bold=True)
        center = Alignment(horizontal='center')

        def write_sheet(sheet, headers, rows, fill):
            for c, lbl in enumerate(headers, 1):
                cell = sheet.cell(1, c, lbl)
                cell.fill = fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(rows, 2):
                for c, value in enumerate(row, 1):
                    cell = sheet.cell(r, c, value)
                    if isinstance(value, (int, float)) and c > 1:
                        cell.number_format = '$ #,##0'

        cierre = get_cierre_mensual(self._mes, self._ano)
        periodo_nomina = period_from_month_year(self._mes, self._ano)

        ws = wb.active
        ws.title = 'Resumen'
        ws['A1'] = 'EXPORTACION COMPLETA CONTABILIDAD MORSA'
        ws['A1'].font = Font(bold=True, size=14)
        summary_rows = [
            ('Periodo visible', f'{MONTHS[self._mes]} {self._ano}'),
            ('Ingresos visibles', cierre['total_ingresos']),
            ('Egresos visibles', cierre['total_egresos']),
            ('Resultado neto visible', cierre['resultado_neto']),
            ('Proveedores totales', len(get_proveedores())),
            ('Egresos totales', len(get_egresos())),
            ('Ingresos totales', len(get_ingresos())),
            ('Nomina periodo visible', periodo_nomina or ''),
            ('Nomina empleados visibles', cierre['nomina']['empleados']),
        ]
        for i, (label, value) in enumerate(summary_rows, start=3):
            ws.cell(i, 1, label)
            ws.cell(i, 2, value)
            if isinstance(value, (int, float)):
                ws.cell(i, 2).number_format = '$ #,##0'
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 22

        proveedores = get_proveedores()
        ws_prov = wb.create_sheet('Proveedores')
        write_sheet(
            ws_prov,
            ['Razon Social', 'NIT', 'Primer Nombre', 'Segundo Nombre', 'Primer Apellido',
             'Segundo Apellido', 'Direccion', 'Telefono', 'Correo'],
            [
                (
                    p['razon_social'], p['nit'], p['primer_nombre'], p['segundo_nombre'],
                    p['primer_apellido'], p['segundo_apellido'], p['direccion'],
                    p['telefono'], p['correo']
                )
                for p in proveedores
            ],
            slate_fill,
        )

        egresos = get_egresos()
        ws_eg = wb.create_sheet('Egresos')
        write_sheet(
            ws_eg,
            ['Fecha', 'N° Doc', 'Consecutivo', 'Proveedor', 'NIT', 'Valor', 'Naturaleza del gasto', 'Factura electrónica', 'Obs', 'Origen', 'Ref', 'Periodo'],
            [
                (
                    e['fecha'], e['no_documento'], e['consecutivo'], e['razon_social'], e['nit'],
                    e['valor'] or 0, e['tipo_gasto'], e.get('factura_electronica', 'NO'), e['observaciones'], e.get('source_module', ''),
                    e.get('source_ref', ''), e.get('source_period', '')
                )
                for e in egresos
            ],
            red_fill,
        )

        ingresos = get_ingresos()
        ws_ing = wb.create_sheet('Ingresos')
        write_sheet(
            ws_ing,
            ['Fecha', 'Caja', 'Bancos', 'Tarjeta CR', 'Total Dia'],
            [
                (
                    i['fecha'], i['caja'] or 0, i['bancos'] or 0, i['tarjeta_cr'] or 0,
                    (i['caja'] or 0) + (i['bancos'] or 0) + (i['tarjeta_cr'] or 0)
                )
                for i in ingresos
            ],
            green_fill,
        )

        nomina = get_nomina_resumen(periodo=periodo_nomina) if periodo_nomina else []
        ws_nom = wb.create_sheet('Nomina')
        write_sheet(
            ws_nom,
            ['Periodo', 'Empleado', 'Cedula', 'Dias Q1', 'Neto Q1', 'Dias Q2', 'Neto Q2', 'Deducciones', 'Total Mes'],
            [
                (
                    n['periodo'], n['empleado'], n['cedula'], n['q1_dias'], n['q1_neto'] or 0,
                    n['q2_dias'], n['q2_neto'] or 0, n['total_deduccion'] or 0, n['total_mes'] or 0
                )
                for n in nomina
            ],
            blue_fill,
        )

        novedades = get_nomina_novedades()
        ws_nov = wb.create_sheet('Novedades')
        write_sheet(
            ws_nov,
            ['Periodo', 'Fecha', 'Empleado', 'Cedula', 'Quincena', 'Naturaleza', 'Tipo', 'Valor', 'Obs', 'Origen'],
            [
                (
                    n['periodo'], n['fecha'], n['empleado'], n['cedula'], n['quincena'],
                    n['naturaleza'], n['tipo_novedad'], n['valor'] or 0,
                    n['observaciones'], n['origen_archivo']
                )
                for n in novedades
            ],
            purple_fill,
        )

        seg = get_nomina_seg_social(periodo=periodo_nomina) if periodo_nomina else []
        ws_seg = wb.create_sheet('SegSocial')
        write_sheet(
            ws_seg,
            ['Periodo', 'Grupo', 'Concepto', 'Valor', 'Obs', 'Origen'],
            [
                (s['periodo'], s['grupo'], s['concepto'], s['valor'] or 0, s['observaciones'], s['origen_archivo'])
                for s in seg
            ],
            purple_fill,
        )

        ws_cierre = wb.create_sheet('Cierre Visible')
        write_sheet(
            ws_cierre,
            ['Concepto', 'Valor'],
            [
                ('Periodo', cierre['periodo'] or f'{MONTHS[self._mes]} {self._ano}'),
                ('Ingresos', cierre['total_ingresos']),
                ('Egresos operativos', cierre['egresos_operativos']),
                ('Nomina empleados', cierre['egresos_nomina']),
                ('Seguridad social', cierre['egresos_seg_social']),
                ('Novedades deduccion', cierre['novedades_deduccion']),
                ('Total egresos', cierre['total_egresos']),
                ('Resultado neto', cierre['resultado_neto']),
            ],
            header_fill,
        )

        try:
            wb.save(path)
        except PermissionError:
            messagebox.showerror('Error', 'No se pudo guardar el archivo. Cierra el Excel si lo tienes abierto e intenta de nuevo.')
            return
        except OSError as exc:
            messagebox.showerror('Error', f'No se pudo guardar el archivo.\n\n{exc}')
            return
        messagebox.showinfo('Exito', f'Exportación completa generada:\n{path}')

    def _export_workbook(self, cierre_only):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror('Error', 'Instala openpyxl: pip install openpyxl')
            return

        suffix = 'Cierre' if cierre_only else 'Reporte'
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialfile=f'{suffix}_{MONTHS[self._mes]}_{self._ano}.xlsx',
        )
        if not path:
            return

        cierre = get_cierre_mensual(self._mes, self._ano)
        periodo_nomina = period_from_month_year(self._mes, self._ano)
        egresos = get_egresos(mes=self._mes, ano=self._ano)
        ingresos = get_ingresos(mes=self._mes, ano=self._ano)
        nomina = get_nomina_resumen(periodo=periodo_nomina)
        novedades = get_nomina_novedades(periodo=periodo_nomina)
        seg = get_nomina_seg_social(periodo=periodo_nomina)
        nomina_stats = get_nomina_stats(periodo=periodo_nomina)

        wb = openpyxl.Workbook()
        header_fill = PatternFill('solid', fgColor='1e3a5f')
        green_fill = PatternFill('solid', fgColor='16a34a')
        red_fill = PatternFill('solid', fgColor='dc2626')
        blue_fill = PatternFill('solid', fgColor='1d4ed8')
        purple_fill = PatternFill('solid', fgColor='7c3aed')
        white_font = Font(color='FFFFFF', bold=True)
        center = Alignment(horizontal='center')

        ws = wb.active
        ws.title = 'Cierre'
        ws['A1'] = f'CIERRE MENSUAL {MONTHS[self._mes].upper()} {self._ano}'
        ws['A1'].font = Font(bold=True, size=14)
        rows = [
            ('Ingresos', cierre['total_ingresos']),
            ('Egresos operativos', cierre['egresos_operativos']),
            ('Nomina empleados', cierre['egresos_nomina']),
            ('Seguridad social', cierre['egresos_seg_social']),
            ('Novedades deduccion', cierre['novedades_deduccion']),
            ('Total egresos', cierre['total_egresos']),
            ('Resultado neto', cierre['resultado_neto']),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws.cell(i, 1, label)
            ws.cell(i, 2, value)
            ws.cell(i, 2).number_format = '$ #,##0'
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18

        if not cierre_only:
            ws_res = wb.create_sheet('Resumen')
            summary = [
                ('Periodo nomina', periodo_nomina or ''),
                ('Empleados', nomina_stats['empleados']),
                ('Total nomina', nomina_stats['total_nomina']),
                ('Novedades +', nomina_stats['total_novedades_devengado']),
                ('Novedades -', nomina_stats['total_novedades_deduccion']),
                ('Seg. social', nomina_stats['total_seg_social']),
                ('Nomina integrada', nomina_stats['total_nomina_integrada']),
            ]
            for i, (label, value) in enumerate(summary, start=1):
                ws_res.cell(i, 1, label)
                ws_res.cell(i, 2, value)
                if isinstance(value, (int, float)):
                    ws_res.cell(i, 2).number_format = '$ #,##0'

            ws_eg = wb.create_sheet('Egresos')
            headers = ['Fecha', 'N° Doc', 'Proveedor', 'NIT', 'Valor', 'Naturaleza del gasto', 'Factura electrónica', 'Origen', 'Obs']
            for c, lbl in enumerate(headers, 1):
                cell = ws_eg.cell(1, c, lbl)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(egresos, 2):
                ws_eg.cell(r, 1, row['fecha'])
                ws_eg.cell(r, 2, row['no_documento'] or '')
                ws_eg.cell(r, 3, row['razon_social'] or '')
                ws_eg.cell(r, 4, row['nit'] or '')
                ws_eg.cell(r, 5, row['valor'] or 0).number_format = '$ #,##0'
                ws_eg.cell(r, 6, row['tipo_gasto'] or '')
                ws_eg.cell(r, 7, row.get('factura_electronica', 'NO') or 'NO')
                ws_eg.cell(r, 8, row['source_module'] or '')
                ws_eg.cell(r, 9, row['observaciones'] or '')

            ws_ing = wb.create_sheet('Ingresos')
            for c, lbl in enumerate(['Fecha', 'Caja', 'Bancos', 'Tarjeta', 'Total'], 1):
                cell = ws_ing.cell(1, c, lbl)
                cell.fill = green_fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(ingresos, 2):
                total = (row['caja'] or 0) + (row['bancos'] or 0) + (row['tarjeta_cr'] or 0)
                ws_ing.cell(r, 1, row['fecha'])
                ws_ing.cell(r, 2, row['caja'] or 0).number_format = '$ #,##0'
                ws_ing.cell(r, 3, row['bancos'] or 0).number_format = '$ #,##0'
                ws_ing.cell(r, 4, row['tarjeta_cr'] or 0).number_format = '$ #,##0'
                ws_ing.cell(r, 5, total).number_format = '$ #,##0'

            ws_nom = wb.create_sheet('Nomina')
            for c, lbl in enumerate(['Empleado', 'Cedula', 'Neto Q1', 'Neto Q2', 'Deducciones', 'Total Mes'], 1):
                cell = ws_nom.cell(1, c, lbl)
                cell.fill = blue_fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(nomina, 2):
                ws_nom.cell(r, 1, row['empleado'])
                ws_nom.cell(r, 2, row['cedula'])
                ws_nom.cell(r, 3, row['q1_neto'] or 0).number_format = '$ #,##0'
                ws_nom.cell(r, 4, row['q2_neto'] or 0).number_format = '$ #,##0'
                ws_nom.cell(r, 5, row['total_deduccion'] or 0).number_format = '$ #,##0'
                ws_nom.cell(r, 6, row['total_mes'] or 0).number_format = '$ #,##0'

            ws_nov = wb.create_sheet('Novedades')
            for c, lbl in enumerate(['Fecha', 'Empleado', 'Quincena', 'Naturaleza', 'Tipo', 'Valor', 'Obs'], 1):
                cell = ws_nov.cell(1, c, lbl)
                cell.fill = red_fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(novedades, 2):
                ws_nov.cell(r, 1, row['fecha'])
                ws_nov.cell(r, 2, row['empleado'])
                ws_nov.cell(r, 3, row['quincena'])
                ws_nov.cell(r, 4, row['naturaleza'])
                ws_nov.cell(r, 5, row['tipo_novedad'])
                ws_nov.cell(r, 6, row['valor'] or 0).number_format = '$ #,##0'
                ws_nov.cell(r, 7, row['observaciones'] or '')

            ws_seg = wb.create_sheet('SegSocial')
            for c, lbl in enumerate(['Grupo', 'Concepto', 'Valor', 'Obs'], 1):
                cell = ws_seg.cell(1, c, lbl)
                cell.fill = purple_fill
                cell.font = white_font
                cell.alignment = center
            for r, row in enumerate(seg, 2):
                ws_seg.cell(r, 1, row['grupo'] or '')
                ws_seg.cell(r, 2, row['concepto'])
                ws_seg.cell(r, 3, row['valor'] or 0).number_format = '$ #,##0'
                ws_seg.cell(r, 4, row['observaciones'] or '')

        try:
            wb.save(path)
        except PermissionError:
            messagebox.showerror('Error', 'No se pudo guardar el archivo. Cierra el Excel si lo tienes abierto e intenta de nuevo.')
            return
        except OSError as exc:
            messagebox.showerror('Error', f'No se pudo guardar el archivo.\n\n{exc}')
            return
        messagebox.showinfo('Exito', f'Archivo exportado:\n{path}')
