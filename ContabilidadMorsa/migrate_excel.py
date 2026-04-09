"""
Ejecutar una sola vez para importar el Excel base a la base de datos activa.
Uso: python migrate_excel.py
"""
import os
import sys
import openpyxl
from datetime import datetime

# Agregar la carpeta raiz al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import init_db, write_transaction

EXCEL_PATH = ''


def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.startswith('='):
        return ''
    return s


def clean_tipo(val):
    t = clean(val).upper()
    return t if t else 'GASTO'


def migrate(path: str | None = None):
    source_path = (path or EXCEL_PATH or '').strip()
    if not source_path or not os.path.exists(source_path):
        print(f'ERROR: No se encontro el archivo:\n  {source_path or "(sin ruta)"}')
        sys.exit(1)

    init_db()

    wb = openpyxl.load_workbook(source_path, data_only=True)
    with write_transaction(checkpoint=True, checkpoint_mode='TRUNCATE') as conn:
        print('Limpiando datos anteriores...')
        conn.execute('DELETE FROM ingresos')
        conn.execute('DELETE FROM egresos')
        conn.execute('DELETE FROM proveedores')

        # ── 1. BASE DE DATOS → proveedores ──────────────────────────────────
        print('Importando proveedores...')
        ws = wb['BASE DE DATOS']
        count_prov = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            razon = clean(row[0])
            if not razon:
                continue
            nit = clean(str(row[1])) if row[1] else ''
            tipo = clean_tipo(row[9]) if len(row) > 9 else ''
            conn.execute(
                'INSERT INTO proveedores '
                '(razon_social, nit, primer_nombre, segundo_nombre, primer_apellido, '
                'segundo_apellido, direccion, telefono, correo, tipo) VALUES (?,?,?,?,?,?,?,?,?,?)',
                (razon, nit, clean(row[2]), clean(row[3]), clean(row[4]),
                 clean(row[5]), clean(row[6]),
                 clean(str(row[7])) if row[7] else '', clean(row[8]), tipo)
            )
            count_prov += 1
        print(f'  → {count_prov} proveedores importados')

        # ── 2. MARZO → egresos ──────────────────────────────────────────────
        print('Importando egresos...')
        ws = wb['MARZO']
        count_eg = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            fecha = row[0]
            if not isinstance(fecha, datetime):
                continue
            valor = row[12]
            if valor is None:
                continue

            razon = clean(row[4])
            if not razon or razon.startswith('='):
                razon = 'SIN PROVEEDOR'

            tipo = clean_tipo(row[13])
            if tipo.startswith('=') or not tipo:
                tipo = 'GASTO'

            nit = clean(str(row[3])) if row[3] and not str(row[3]).startswith('=') else ''
            no_doc = clean(str(row[1])) if row[1] else ''
            consec = clean(str(row[2])) if row[2] else ''

            conn.execute(
                'INSERT INTO egresos (fecha, no_documento, consecutivo, razon_social, nit, valor, tipo_gasto, factura_electronica) '
                'VALUES (?,?,?,?,?,?,?,?)',
                (fecha.strftime('%Y-%m-%d'), no_doc, consec, razon, nit, float(valor), tipo, 'NO')
            )
            count_eg += 1
        print(f'  → {count_eg} egresos importados')

        # ── 3. INGRESOS → ingresos ──────────────────────────────────────────
        print('Importando ingresos...')
        ws = wb['INGRESOS']
        count_ing = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            fecha = row[0]
            if not isinstance(fecha, datetime):
                continue
            caja = float(row[1]) if isinstance(row[1], (int, float)) else 0.0
            bancos = float(row[2]) if isinstance(row[2], (int, float)) else 0.0
            tarjeta = float(row[3]) if isinstance(row[3], (int, float)) else 0.0
            if caja == 0 and bancos == 0 and tarjeta == 0:
                continue
            conn.execute(
                'INSERT OR REPLACE INTO ingresos (fecha, caja, bancos, tarjeta_cr) VALUES (?,?,?,?)',
                (fecha.strftime('%Y-%m-%d'), caja, bancos, tarjeta)
            )
            count_ing += 1
        print(f'  → {count_ing} registros de ingresos importados')

    print('\n✓ Migracion completada exitosamente.')


if __name__ == '__main__':
    migrate(sys.argv[1] if len(sys.argv) > 1 else None)
