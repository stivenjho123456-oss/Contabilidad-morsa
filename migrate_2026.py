"""
Importa los datos de Enero, Febrero y Marzo 2026 a la base de datos.
Solo agrega registros nuevos — NO borra nada existente.
Uso: python3 migrate_2026.py
"""
import os
import sys
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(ROOT, 'ContabilidadMorsa'))

from database import get_connection, init_db  # noqa: E402


def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s


def clean_tipo(val):
    t = clean(val).upper()
    return t if t else 'GASTO'


def to_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str) and len(val) == 10:
        return val
    return None


def import_proveedores(conn, ws):
    """Inserta proveedores que NO existen ya (match por razon_social)."""
    existing = {
        r[0].strip().upper()
        for r in conn.execute('SELECT razon_social FROM proveedores').fetchall()
    }
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        razon = clean(row[0])
        if not razon or razon.upper() in existing:
            continue
        nit  = clean(str(row[1])) if row[1] else ''
        tipo = clean_tipo(row[9]) if len(row) > 9 else ''
        conn.execute(
            'INSERT INTO proveedores '
            '(razon_social, nit, primer_nombre, segundo_nombre, primer_apellido, '
            'segundo_apellido, direccion, telefono, correo, tipo) VALUES (?,?,?,?,?,?,?,?,?,?)',
            (razon, nit, clean(row[2]), clean(row[3]), clean(row[4]),
             clean(row[5]), clean(row[6]),
             clean(str(row[7])) if row[7] else '', clean(row[8]), tipo)
        )
        existing.add(razon.upper())
        added += 1
    conn.commit()
    return added


def import_egresos(conn, ws, sheet_name):
    """
    Inserta egresos que NO existan ya (match por fecha+razon_social+valor).
    Columnas del Excel: 0=fecha 1=no_doc 2=consec 3=nit 4=razon_social 12=valor 13=tipo_gasto
    """
    existing = set(
        conn.execute(
            "SELECT fecha, UPPER(TRIM(razon_social)), ROUND(valor,2) FROM egresos"
        ).fetchall()
    )
    added = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        fecha = to_date(row[0])
        if not fecha:
            continue
        valor = row[12]
        if valor is None:
            continue
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            continue
        if valor <= 0:
            continue

        razon = clean(row[4]) or 'SIN PROVEEDOR'
        key = (fecha, razon.upper().strip(), round(valor, 2))
        if key in existing:
            skipped += 1
            continue

        tipo  = clean_tipo(row[13])
        nit   = clean(str(row[3])) if row[3] and not str(row[3]).startswith('=') else ''
        no_doc = clean(str(row[1])) if row[1] else ''
        consec = clean(str(row[2])) if row[2] else ''

        conn.execute(
            'INSERT INTO egresos '
            '(fecha, no_documento, consecutivo, razon_social, nit, valor, tipo_gasto, factura_electronica) '
            'VALUES (?,?,?,?,?,?,?,?)',
            (fecha, no_doc, consec, razon, nit, valor, tipo, 'NO')
        )
        existing.add(key)
        added += 1
    conn.commit()
    return added, skipped


def import_ingresos_sheet(conn, ws, start_row=4):
    """
    Inserta ingresos que NO existan ya (match por fecha).
    Evita duplicados por fecha (un registro por día).
    """
    existing = {
        r[0] for r in conn.execute('SELECT fecha FROM ingresos').fetchall()
    }
    added = skipped = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        fecha = to_date(row[0])
        if not fecha:
            continue
        caja    = float(row[1]) if isinstance(row[1], (int, float)) else 0.0
        bancos  = float(row[2]) if isinstance(row[2], (int, float)) else 0.0
        tarjeta = float(row[3]) if isinstance(row[3], (int, float)) else 0.0
        if caja == 0 and bancos == 0 and tarjeta == 0:
            continue
        if fecha in existing:
            skipped += 1
            continue
        conn.execute(
            'INSERT INTO ingresos (fecha, caja, bancos, tarjeta_cr) VALUES (?,?,?,?)',
            (fecha, caja, bancos, tarjeta)
        )
        existing.add(fecha)
        added += 1
    conn.commit()
    return added, skipped


def main():
    init_db()
    conn = get_connection()

    files = [
        ('ENERO 2026.xlsx',    'ENERO',    'INGRESOS', 'BASE DE DATOS'),
        ('FEBRERO 2026.xlsx',  'FEBRERO',  'INGRESOS', 'BASE DE DATOS'),
    ]
    ingresos_extra = [
        ('Ingresos_3_2026.xlsx', 'Ingresos', 4),
    ]

    for fname, sheet_eg, sheet_ing, sheet_bd in files:
        path = os.path.join(ROOT, fname)
        if not os.path.exists(path):
            print(f'⚠  Archivo no encontrado: {fname}  (se omite)')
            continue

        print(f'\n📂  {fname}')
        wb = openpyxl.load_workbook(path, data_only=True)

        # Proveedores
        if sheet_bd in wb.sheetnames:
            n = import_proveedores(conn, wb[sheet_bd])
            print(f'   Proveedores nuevos:  {n}')

        # Egresos
        if sheet_eg in wb.sheetnames:
            added, skipped = import_egresos(conn, wb[sheet_eg], sheet_eg)
            print(f'   Egresos agregados:   {added}  (ya existían: {skipped})')

        # Ingresos
        if sheet_ing in wb.sheetnames:
            added, skipped = import_ingresos_sheet(conn, wb[sheet_ing])
            print(f'   Ingresos agregados:  {added}  (ya existían: {skipped})')

    # Marzo 2026 ingresos (archivo separado)
    for fname, sheet, start_row in ingresos_extra:
        path = os.path.join(ROOT, fname)
        if not os.path.exists(path):
            print(f'\n⚠  Archivo no encontrado: {fname}  (se omite)')
            continue
        print(f'\n📂  {fname}')
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet in wb.sheetnames:
            added, skipped = import_ingresos_sheet(conn, wb[sheet], start_row=start_row)
            print(f'   Ingresos Marzo 2026: {added} agregados  ({skipped} ya existían)')

    conn.close()

    # Resumen final
    conn2 = get_connection()
    print('\n── Resumen final de la base de datos ──')
    for table in ['egresos', 'ingresos', 'proveedores']:
        c = conn2.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
        print(f'   {table:<15} {c} registros')

    print('\n── Egresos por mes ──')
    rows = conn2.execute(
        "SELECT strftime('%Y-%m', fecha) mes, COUNT(*) FROM egresos GROUP BY mes ORDER BY mes"
    ).fetchall()
    for r in rows:
        print(f'   {r[0]}  →  {r[1]} egresos')

    print('\n── Ingresos por mes ──')
    rows = conn2.execute(
        "SELECT strftime('%Y-%m', fecha) mes, COUNT(*) FROM ingresos GROUP BY mes ORDER BY mes"
    ).fetchall()
    for r in rows:
        print(f'   {r[0]}  →  {r[1]} ingresos')

    conn2.close()
    print('\n✓ Migración 2026 completada.')


if __name__ == '__main__':
    main()
