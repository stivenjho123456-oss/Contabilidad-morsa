from pathlib import Path
import logging
import os
import secrets
import sys
import sqlite3
import tempfile
import threading
import time
from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.staticfiles import StaticFiles
from starlette.exceptions import HTTPException as StarletteHTTPException
from pydantic import BaseModel, Field
from starlette.status import HTTP_422_UNPROCESSABLE_CONTENT, HTTP_500_INTERNAL_SERVER_ERROR
from starlette.requests import Request

ROOT_DIR = Path(__file__).resolve().parents[3]
CORE_APP_DIR = ROOT_DIR / "ContabilidadMorsa"
FRONTEND_DIST_DIR = ROOT_DIR / "apps" / "frontend" / "dist"
if str(CORE_APP_DIR) not in sys.path:
    sys.path.insert(0, str(CORE_APP_DIR))

# Asegura que el directorio de este archivo esté en el path (necesario en Render)
_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

# ── Adaptador de base de datos (cloud-first) ─────────────────────────────────
from db_adapter import (  # noqa: E402
    ALLOW_SQLITE_FALLBACK,
    USE_POSTGRES,
    DatabaseSchemaError,
    get_pg_database_health,
    get_pg_schema_report,
    require_pg_schema,
)
if USE_POSTGRES:
    # Parchamos get_connection en database para que use PostgreSQL
    import db_adapter as _db_adapter  # noqa: E402
    import database as _db_module      # noqa: E402
    _db_module.get_connection = _db_adapter.get_pg_connection
    logger_pre = logging.getLogger("contabilidad_morsa.boot")
    logger_pre.info("Modo PostgreSQL activado (DATABASE_URL presente)")

from app_paths import get_log_dir  # noqa: E402

from database import (  # noqa: E402
    AppValidationError,
    auth_bootstrap_required,
    create_archivo_blob,
    create_caja_ajuste,
    delete_archivo_blob,
    delete_cuadre_caja,
    delete_egreso,
    delete_ingreso,
    delete_nomina_asistencia,
    delete_nomina_novedad,
    delete_proveedor,
    get_auditoria,
    get_archivo_blob,
    get_caja_ajustes,
    get_caja_apertura_context,
    get_connection,
    get_cierre_mensual,
    get_caja_movimientos_detalle,
    get_caja_snapshot_by_fecha,
    get_cuadres_caja,
    get_cuadre_caja_by_fecha,
    get_database_health,
    get_dashboard_stats,
    get_egresos,
    get_ingresos,
    get_nomina_bundle,
    get_nomina_novedades,
    get_nomina_resumen,
    get_nomina_seg_social,
    get_proveedores,
    get_saldo_inicial_sugerido,
    get_tipos_gasto_distintos,
    init_db,
    is_period_closed,
    list_cierres_mensuales,
    log_auditoria,
    month_year_from_date,
    period_from_month_year,
    save_cuadre_caja,
    save_nomina_asistencia,
    save_nomina_novedad,
    save_egreso,
    save_ingreso,
    save_proveedor,
    set_cierre_mensual,
    sync_nomina_to_egresos,
)
from auth_service import (  # noqa: E402
    SESSION_HEADER as API_TOKEN_HEADER,
    SESSION_SCHEME as API_TOKEN_SCHEME,
    auth_status as get_auth_status,
    authenticate_user,
    bootstrap_admin_env_configured,
    bootstrap_admin_account,
    ensure_bootstrap_admin_from_env,
    resolve_session,
    revoke_session,
)
import migrate_excel as migrate_excel_module  # noqa: E402
import migrate_nomina as migrate_nomina_module  # noqa: E402


LOG_DIR = get_log_dir()
LOG_FILE = LOG_DIR / "backend.log"
ENABLE_DOCS = os.getenv("MORSA_ENABLE_DOCS") == "1"
DEFAULT_ALLOWED_ORIGINS = [
    "http://127.0.0.1:5175",
    "http://localhost:5175",
    "http://127.0.0.1:8010",
    "http://localhost:8010",
    "https://contabilidad-morsa.vercel.app",
]
ALLOWED_SUPPORT_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
ALLOWED_SUPPORT_CONTENT_TYPES = {"application/pdf", "image/png", "image/jpeg", "image/webp"}
ALLOWED_IMPORT_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_SUPPORT_SIZE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_SIZE_BYTES = 25 * 1024 * 1024
UPLOAD_CHUNK_SIZE = 1024 * 1024
SYSTEM_SUMMARY_CACHE_SECONDS = 20
RUNTIME_STATUS_CACHE_SECONDS = 30
RUNTIME_QUERY_CACHE_SECONDS = 20
_RUNTIME_CACHE_LOCK = threading.RLock()
LOG_HANDLERS = [logging.StreamHandler()]
try:
    LOG_HANDLERS.insert(0, logging.FileHandler(LOG_FILE, encoding="utf-8"))
except OSError:
    pass
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    handlers=LOG_HANDLERS,
)
logger = logging.getLogger("contabilidad_morsa.backend")

app = FastAPI(
    title="Contabilidad Morsa API",
    version="1.0.0",
    docs_url="/docs" if ENABLE_DOCS else None,
    redoc_url=None,
    openapi_url="/openapi.json" if ENABLE_DOCS else None,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in os.getenv("MORSA_ALLOWED_ORIGINS", ",".join(DEFAULT_ALLOWED_ORIGINS)).split(",") if origin.strip()],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", API_TOKEN_HEADER],
)


class ProveedorPayload(BaseModel):
    razon_social: str
    nit: str = ""
    primer_nombre: str = ""
    segundo_nombre: str = ""
    primer_apellido: str = ""
    segundo_apellido: str = ""
    direccion: str = ""
    telefono: str = ""
    correo: str = ""


class IngresoPayload(BaseModel):
    fecha: str
    caja: float = 0
    bancos: float = 0
    tarjeta_cr: float = 0


class EgresoPayload(BaseModel):
    fecha: str
    no_documento: str = ""
    consecutivo: str = ""
    proveedor_id: int | None = None
    razon_social: str
    nit: str = ""
    valor: float = Field(gt=0)
    tipo_gasto: str
    canal_pago: str = "Otro"
    factura_electronica: str = "NO"
    observaciones: str = ""
    source_module: str = ""
    source_ref: str = ""
    source_period: str = ""


class CuadreCajaPayload(BaseModel):
    fecha: str
    saldo_inicial: float = 0
    saldo_real: float | None = None
    observaciones: str = ""


class CajaAjustePayload(BaseModel):
    fecha: str
    tipo: str
    valor: float = Field(gt=0)
    motivo: str
    observaciones: str = ""


class NovedadPayload(BaseModel):
    periodo: str
    fecha: str
    empleado: str
    cedula: str = ""
    quincena: str = ""
    naturaleza: str
    tipo_novedad: str
    valor: float = Field(gt=0)
    observaciones: str = ""
    origen_archivo: str = ""


class AsistenciaPayload(BaseModel):
    periodo: str
    empleado: str
    cedula: str = ""
    dia: int = Field(ge=1, le=31)
    quincena: str = ""
    estado: str
    origen_archivo: str = ""


class SyncNominaPayload(BaseModel):
    periodo: str | None = None


class CierreMensualPayload(BaseModel):
    mes: int
    ano: int
    observacion: str = ""


class AuthLoginPayload(BaseModel):
    username: str
    password: str


class AuthBootstrapPayload(BaseModel):
    username: str
    full_name: str
    password: str
    password_confirm: str


def _handle_validation(exc: Exception):
    if isinstance(exc, AppValidationError):
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    raise exc


def _api_ok(data: Any = None, message: str | None = None):
    return {"ok": True, "message": message, "data": data}


def _sanitize_filename(value: str) -> str:
    raw = Path(value or "soporte").name
    cleaned = "".join(ch if ch.isalnum() or ch in {".", "-", "_"} else "_" for ch in raw)
    return cleaned or "soporte"


def _startup_state():
    return {
        "db_health": None,
        "db_health_checked_at": 0.0,
        "schema_status": None,
        "schema_status_checked_at": 0.0,
        "system_summary_cache": None,
        "system_summary_cache_at": 0.0,
        "query_cache": {},
        "query_inflight": {},
        "query_cache_version": 0,
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "log_file": str(LOG_FILE),
        "deployment_mode": "cloud",
    }


def _current_db_health():
    if USE_POSTGRES:
        return get_pg_database_health()
    return get_database_health()


def _current_schema_status():
    if USE_POSTGRES:
        return get_pg_schema_report()
    return None


def _refresh_runtime_status(runtime: dict[str, Any], *, force: bool = False):
    now = time.monotonic()
    if force or runtime.get("db_health") is None or (now - runtime.get("db_health_checked_at", 0.0)) >= RUNTIME_STATUS_CACHE_SECONDS:
        runtime["db_health"] = _current_db_health()
        runtime["db_health_checked_at"] = now
    if force or runtime.get("schema_status") is None or (now - runtime.get("schema_status_checked_at", 0.0)) >= RUNTIME_STATUS_CACHE_SECONDS:
        runtime["schema_status"] = _current_schema_status()
        runtime["schema_status_checked_at"] = now
    return runtime["db_health"], runtime["schema_status"]


def _runtime_state():
    runtime = getattr(app.state, "runtime", None)
    if runtime is None:
        runtime = _startup_state()
        app.state.runtime = runtime
    return runtime


def _invalidate_runtime_caches():
    runtime = _runtime_state()
    with _RUNTIME_CACHE_LOCK:
        runtime["query_cache"].clear()
        runtime["query_cache_version"] = runtime.get("query_cache_version", 0) + 1
        runtime["system_summary_cache"] = None
        runtime["system_summary_cache_at"] = 0.0


def _cached_runtime_query(cache_key: tuple[Any, ...], loader, ttl_seconds: int = RUNTIME_QUERY_CACHE_SECONDS):
    runtime = _runtime_state()
    cache_version = None
    while True:
        with _RUNTIME_CACHE_LOCK:
            current_version = runtime.get("query_cache_version", 0)
            entry = runtime["query_cache"].get(cache_key)
            if entry and entry["version"] == current_version and (time.monotonic() - entry["stored_at"]) < ttl_seconds:
                return entry["value"]
            inflight = runtime["query_inflight"].get(cache_key)
            if inflight is None:
                inflight = threading.Event()
                runtime["query_inflight"][cache_key] = inflight
                is_loader = True
                cache_version = current_version
            else:
                is_loader = False
        if is_loader:
            break
        inflight.wait()

    try:
        value = loader()
    except Exception:
        with _RUNTIME_CACHE_LOCK:
            pending = runtime["query_inflight"].pop(cache_key, None)
            if pending is not None:
                pending.set()
        raise

    with _RUNTIME_CACHE_LOCK:
        if runtime.get("query_cache_version", 0) == cache_version:
            runtime["query_cache"][cache_key] = {
                "value": value,
                "stored_at": time.monotonic(),
                "version": cache_version,
            }
        pending = runtime["query_inflight"].pop(cache_key, None)
        if pending is not None:
            pending.set()
    return value


def _apply_security_headers(response, *, is_api: bool):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; "
        "script-src 'self'; connect-src 'self' https:; frame-ancestors 'none'; base-uri 'self'; form-action 'self'",
    )
    if is_api:
        response.headers.setdefault("Cache-Control", "no-store")
    return response


def _public_request_path(path: str):
    return (
        path == "/"
        or path == "/health"
        or path == "/api/auth/status"
        or path == "/api/auth/login"
        or path == "/api/auth/bootstrap"
        or (ENABLE_DOCS and path in {"/docs", "/openapi.json"})
        or path == "/favicon.svg"
        or path == "/icons.svg"
        or path.startswith("/assets/")
    )


def _parse_bearer_token(header_value: str | None):
    raw = (header_value or "").strip()
    if not raw:
        return ""
    parts = raw.split(None, 1)
    if len(parts) == 2 and parts[0].lower() == API_TOKEN_SCHEME.lower():
        return parts[1].strip()
    return ""


def _client_ip(request: Request):
    forwarded = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    if forwarded:
        return forwarded
    return request.client.host if request.client else ""


def _serialize_db_health(info: dict[str, Any] | None):
    if not info:
        return None
    return {
        "backend": info.get("backend", "postgresql"),
        "exists": info.get("exists", False),
        "size_bytes": info.get("size_bytes", 0),
        "ok": info.get("ok", False),
        "connected": info.get("connected"),
        "database": info.get("database"),
        "server_version": info.get("server_version"),
        "integrity": info.get("integrity"),
        "error": info.get("error"),
    }


def _serialize_schema_status(info: dict[str, Any] | None):
    if not info:
        return None
    return {
        "backend": info.get("backend", "postgresql"),
        "ok": info.get("ok", False),
        "tables_checked": info.get("tables_checked", 0),
        "missing_tables": info.get("missing_tables", []),
        "missing_columns": info.get("missing_columns", {}),
        "error": info.get("error"),
    }


def _serialize_egreso(row: dict[str, Any]):
    data = dict(row)
    data["has_support"] = bool(data.get("support_file_id") or data.get("soporte_name"))
    data.pop("soporte_path", None)
    return data


def _serialize_audit_entry(row: dict[str, Any]):
    data = dict(row)
    data.pop("snapshot", None)
    return data


def _sanitize_excel_value(value: Any):
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    return value


def _validate_uploaded_filename(upload: UploadFile, allowed_extensions: set[str], allowed_content_types: set[str] | None = None):
    if not upload.filename:
        raise HTTPException(status_code=400, detail="Debes seleccionar un archivo.")
    filename = _sanitize_filename(upload.filename)
    suffix = Path(filename).suffix.lower()
    if suffix not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Tipo de archivo no permitido: {suffix or 'sin extensión'}.")
    if allowed_content_types and upload.content_type:
        content_type = upload.content_type.lower()
        if content_type not in allowed_content_types:
            raise HTTPException(status_code=400, detail="El tipo MIME del archivo no está permitido.")
    return filename, suffix


async def _read_uploaded_binary(
    upload: UploadFile,
    *,
    allowed_extensions: set[str],
    allowed_content_types: set[str] | None,
    max_size_bytes: int,
):
    filename, _ = _validate_uploaded_filename(upload, allowed_extensions, allowed_content_types)
    chunks: list[bytes] = []
    total_size = 0
    try:
        while True:
            chunk = await upload.read(UPLOAD_CHUNK_SIZE)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > max_size_bytes:
                raise HTTPException(status_code=400, detail=f"El archivo supera el límite de {max_size_bytes // (1024 * 1024)} MB.")
            chunks.append(chunk)
    finally:
        await upload.close()
    if total_size == 0:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    return b"".join(chunks), filename, total_size, (upload.content_type or "application/octet-stream").lower()


async def _persist_uploaded_file(
    upload: UploadFile,
    *,
    destination_dir: Path,
    allowed_extensions: set[str],
    allowed_content_types: set[str] | None,
    max_size_bytes: int,
    prefix: str,
):
    filename, _ = _validate_uploaded_filename(upload, allowed_extensions, allowed_content_types)
    destination_dir.mkdir(parents=True, exist_ok=True)
    target = destination_dir / f"{prefix}_{filename}"
    total_size = 0
    try:
        with target.open("wb") as fh:
            while True:
                chunk = await upload.read(UPLOAD_CHUNK_SIZE)
                if not chunk:
                    break
                total_size += len(chunk)
                if total_size > max_size_bytes:
                    raise HTTPException(status_code=400, detail=f"El archivo supera el límite de {max_size_bytes // (1024 * 1024)} MB.")
                fh.write(chunk)
    except Exception:
        target.unlink(missing_ok=True)
        raise
    finally:
        await upload.close()
    if total_size == 0:
        target.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    return target, filename, total_size


async def _save_import_upload(upload: UploadFile, import_kind: str):
    temp_dir = Path(tempfile.gettempdir()) / "contabilidad-morsa-imports" / import_kind
    prefix = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(6)}"
    target, filename, total_size = await _persist_uploaded_file(
        upload,
        destination_dir=temp_dir,
        allowed_extensions=ALLOWED_IMPORT_EXTENSIONS,
        allowed_content_types=None,
        max_size_bytes=MAX_IMPORT_SIZE_BYTES,
        prefix=prefix,
    )
    return target, filename, total_size


def _frontend_index_response():
    index_file = FRONTEND_DIST_DIR / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    raise HTTPException(status_code=404, detail="Frontend no compilado.")


def _fetch_one(table: str, row_id: int):
    conn = get_connection()
    try:
        row = conn.execute(f"SELECT * FROM {table} WHERE id=?", (row_id,)).fetchone()
    finally:
        conn.close()
    return dict(row) if row else None


def _require_existing(table: str, row_id: int, label: str):
    row = _fetch_one(table, row_id)
    if not row:
        raise HTTPException(status_code=404, detail=f"{label} no encontrado.")
    return row


def _validate_period(mes: int | None, ano: int | None):
    if mes is None and ano is None:
        return
    if mes is None or ano is None:
        raise HTTPException(status_code=400, detail="Debes enviar mes y año juntos.")
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="El mes debe estar entre 1 y 12.")
    if ano < 2020 or ano > 2100:
        raise HTTPException(status_code=400, detail="El año está fuera de rango.")


def _safe_local_path(value: str | None, fallback: str) -> Path:
    candidate = Path((value or fallback)).expanduser().resolve()
    if not candidate.exists():
        raise HTTPException(status_code=404, detail=f"No existe el archivo: {candidate}")
    if candidate.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel válido.")
    return candidate


def _system_counts():
    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT
                (SELECT COUNT(*) FROM proveedores) AS proveedores,
                (SELECT COUNT(*) FROM egresos) AS egresos,
                (SELECT COUNT(*) FROM ingresos) AS ingresos,
                (SELECT COUNT(*) FROM nomina_resumen) AS nomina_resumen,
                (SELECT COUNT(*) FROM nomina_seg_social) AS nomina_seg_social,
                (SELECT COUNT(*) FROM nomina_novedades) AS nomina_novedades,
                (SELECT COUNT(*) FROM nomina_asistencia) AS nomina_asistencia
            """
        ).fetchone()
        return {
            "proveedores": row["proveedores"],
            "egresos": row["egresos"],
            "ingresos": row["ingresos"],
            "nomina_resumen": row["nomina_resumen"],
            "nomina_seg_social": row["nomina_seg_social"],
            "nomina_novedades": row["nomina_novedades"],
            "nomina_asistencia": row["nomina_asistencia"],
        }
    finally:
        conn.close()


@app.middleware("http")
async def auth_and_security_middleware(request: Request, call_next):
    path = request.url.path
    is_api = path.startswith("/api/")
    mutating_api = is_api and request.method in {"POST", "PUT", "PATCH", "DELETE"} and not path.startswith("/api/auth/")
    if request.method != "OPTIONS" and is_api and not _public_request_path(path):
        incoming_token = _parse_bearer_token(request.headers.get(API_TOKEN_HEADER, ""))
        session = resolve_session(incoming_token)
        if not session:
            return _apply_security_headers(
                JSONResponse(status_code=401, content={"ok": False, "detail": "Sesión inválida o vencida."}),
                is_api=True,
            )
        request.state.current_user = session["user"]
        request.state.auth_session = session
    response = await call_next(request)
    if mutating_api and response.status_code < 400:
        _invalidate_runtime_caches()
    return _apply_security_headers(response, is_api=is_api)


@app.on_event("startup")
def on_startup():
    app.state.runtime = _startup_state()
    if not USE_POSTGRES and not ALLOW_SQLITE_FALLBACK:
        raise RuntimeError("Contabilidad Morsa quedó en modo cloud-only. Debes definir DATABASE_URL o PG_HOST/PG_PASSWORD.")

    try:
        if USE_POSTGRES:
            app.state.runtime["schema_status"] = require_pg_schema()
            if auth_bootstrap_required() and not bootstrap_admin_env_configured():
                raise RuntimeError(
                    "El despliegue cloud requiere un administrador inicial por entorno. "
                    "Define MORSA_ADMIN_USERNAME, MORSA_ADMIN_PASSWORD y MORSA_ADMIN_FULL_NAME antes del primer arranque."
                )
        else:
            init_db()
        ensure_bootstrap_admin_from_env()
        _refresh_runtime_status(app.state.runtime, force=True)
    except DatabaseSchemaError as exc:
        logger.error("Esquema PostgreSQL inválido en startup: %s", exc)
        raise RuntimeError(str(exc)) from exc
    except sqlite3.DatabaseError as exc:
        logger.exception("Fallo de base de datos en startup")
        raise RuntimeError("No fue posible iniciar la base de datos.") from exc
    except Exception as exc:
        logger.exception("Fallo general en startup")
        raise RuntimeError("No fue posible iniciar la aplicación.") from exc


@app.get("/api/auth/status")
def auth_status():
    return _api_ok(get_auth_status())


@app.post("/api/auth/bootstrap")
def auth_bootstrap(payload: AuthBootstrapPayload, request: Request):
    if payload.password != payload.password_confirm:
        raise HTTPException(status_code=400, detail="La confirmación de contraseña no coincide.")
    try:
        session = bootstrap_admin_account(
            payload.username,
            payload.full_name,
            payload.password,
            user_agent=request.headers.get("user-agent", ""),
            ip_address=_client_ip(request),
        )
        return _api_ok(session, message="Cuenta inicial creada correctamente.")
    except Exception as exc:
        _handle_validation(exc)


@app.post("/api/auth/login")
def auth_login(payload: AuthLoginPayload, request: Request):
    session = authenticate_user(
        payload.username,
        payload.password,
        user_agent=request.headers.get("user-agent", ""),
        ip_address=_client_ip(request),
    )
    if not session:
        raise HTTPException(status_code=401, detail="Credenciales inválidas.")
    return _api_ok(session, message="Sesión iniciada.")


@app.get("/api/auth/session")
def auth_session(request: Request):
    session = getattr(request.state, "auth_session", None)
    if not session:
        raise HTTPException(status_code=401, detail="Sesión inválida o vencida.")
    return _api_ok(
        {
            "header": API_TOKEN_HEADER,
            "scheme": API_TOKEN_SCHEME,
            "expires_at": session["expires_at"],
            "user": session["user"],
        }
    )


@app.post("/api/auth/logout")
def auth_logout(request: Request):
    revoke_session(_parse_bearer_token(request.headers.get(API_TOKEN_HEADER, "")))
    return _api_ok(message="Sesión cerrada correctamente.")


@app.exception_handler(HTTPException)
async def fastapi_http_exception_handler(_, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"ok": False, "detail": exc.detail},
    )


@app.exception_handler(StarletteHTTPException)
async def starlette_http_exception_handler(_, exc: StarletteHTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"ok": False, "detail": exc.detail},
    )


@app.exception_handler(RequestValidationError)
async def request_validation_handler(_, exc: RequestValidationError):
    return JSONResponse(
        status_code=HTTP_422_UNPROCESSABLE_CONTENT,
        content={"ok": False, "detail": "Solicitud inválida.", "errors": exc.errors()},
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(_, exc: Exception):
    logger.exception("Error no controlado", exc_info=exc)
    return JSONResponse(
        status_code=HTTP_500_INTERNAL_SERVER_ERROR,
        content={"ok": False, "detail": "Error interno del servidor.", "error_type": exc.__class__.__name__},
    )


@app.get("/health")
def health():
    runtime = getattr(app.state, "runtime", _startup_state())
    db_health = _current_db_health()
    schema_status = _current_schema_status()
    runtime["db_health"] = db_health
    runtime["schema_status"] = schema_status
    is_healthy = bool(db_health.get("ok")) and (schema_status is None or bool(schema_status.get("ok")))
    payload = {
        "status": "ok" if db_health["ok"] else "degraded",
        "service": "contabilidad-morsa-api",
        "version": app.version,
        "db_health": _serialize_db_health(db_health),
        "schema_status": _serialize_schema_status(schema_status),
        "deployment_mode": runtime.get("deployment_mode", "cloud"),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }
    return JSONResponse(status_code=200 if is_healthy else 503, content=payload)


@app.get("/api/system/summary")
def system_summary():
    runtime = getattr(app.state, "runtime", _startup_state())
    db_health, schema_status = _refresh_runtime_status(runtime)
    now = time.monotonic()
    cached = runtime.get("system_summary_cache")
    if not cached or (now - runtime.get("system_summary_cache_at", 0.0)) >= SYSTEM_SUMMARY_CACHE_SECONDS:
        cached = {
            "counts": _system_counts(),
            "cierres": list_cierres_mensuales()[:24],
            "deployment_mode": runtime.get("deployment_mode", "cloud"),
            "storage_mode": "database",
        }
        runtime["system_summary_cache"] = cached
        runtime["system_summary_cache_at"] = now
    return _api_ok(
        {
            "db_health": _serialize_db_health(db_health),
            "schema_status": _serialize_schema_status(schema_status),
            **cached,
        }
    )


@app.get("/", include_in_schema=False)
def root(request: Request):
    accepts = request.headers.get("accept", "")
    if FRONTEND_DIST_DIR.exists() and "text/html" in accepts:
        return _frontend_index_response()
    return _api_ok(
        {
            "service": "Contabilidad Morsa API",
            "version": app.version,
            "docs": "/docs",
            "health": "/health",
        }
    )


@app.get("/api/dashboard")
def dashboard(
    mes: int | None = Query(default=None),
    ano: int | None = Query(default=None),
    include_cierre: bool = Query(default=False),
):
    _validate_period(mes, ano)
    payload = _cached_runtime_query(
        ("dashboard", mes, ano, include_cierre),
        lambda: {
            "stats": get_dashboard_stats(mes=mes, ano=ano),
            **({"cierre": get_cierre_mensual(mes, ano)} if include_cierre and mes and ano else {}),
        },
    )
    return _api_ok(payload)


@app.get("/api/proveedores")
def proveedores(search: str = ""):
    normalized_search = search.strip().lower()
    return _api_ok(
        _cached_runtime_query(
            ("proveedores", normalized_search),
            lambda: get_proveedores(search),
        )
    )


@app.get("/api/proveedores/{prov_id}")
def proveedor_detail(prov_id: int):
    return _api_ok(_require_existing("proveedores", prov_id, "Proveedor"))


@app.post("/api/proveedores")
def create_proveedor(payload: ProveedorPayload):
    try:
        save_proveedor(payload.model_dump())
    except Exception as exc:
        _handle_validation(exc)
    return _api_ok(message="Proveedor creado.")


@app.put("/api/proveedores/{prov_id}")
def update_proveedor(prov_id: int, payload: ProveedorPayload):
    _require_existing("proveedores", prov_id, "Proveedor")
    try:
        save_proveedor(payload.model_dump(), prov_id=prov_id)
    except Exception as exc:
        _handle_validation(exc)
    return _api_ok(message="Proveedor actualizado.")


@app.delete("/api/proveedores/{prov_id}")
def remove_proveedor(prov_id: int):
    _require_existing("proveedores", prov_id, "Proveedor")
    try:
        delete_proveedor(prov_id)
    except Exception as exc:
        _handle_validation(exc)
    return _api_ok(message="Proveedor eliminado.")


@app.get("/api/ingresos")
def ingresos(
    mes: int | None = Query(default=None),
    ano: int | None = Query(default=None),
):
    _validate_period(mes, ano)
    return _api_ok(
        _cached_runtime_query(
            ("ingresos", mes, ano),
            lambda: get_ingresos(mes=mes, ano=ano),
        )
    )


@app.get("/api/ingresos/analisis")
def ingresos_analisis():
    """Totales por canal agrupados por mes — para el panel de análisis de ingresos."""
    def _load_analisis():
        conn = get_connection()
        try:
            rows = conn.execute("""
                SELECT
                    strftime('%Y-%m', fecha) AS mes,
                    COUNT(*)                AS dias,
                    COALESCE(SUM(caja), 0)       AS caja,
                    COALESCE(SUM(bancos), 0)     AS bancos,
                    COALESCE(SUM(tarjeta_cr), 0) AS tarjeta_cr,
                    COALESCE(SUM(caja + bancos + tarjeta_cr), 0) AS total
                FROM ingresos
                WHERE fecha IS NOT NULL
                GROUP BY mes
                ORDER BY mes
            """).fetchall()
            meses = [dict(r) for r in rows]

            totales = conn.execute("""
                SELECT
                    COALESCE(SUM(caja), 0)       AS caja,
                    COALESCE(SUM(bancos), 0)     AS bancos,
                    COALESCE(SUM(tarjeta_cr), 0) AS tarjeta_cr,
                    COALESCE(SUM(caja + bancos + tarjeta_cr), 0) AS total,
                    COUNT(DISTINCT strftime('%Y-%m', fecha))      AS meses_con_datos
                FROM ingresos
                WHERE fecha IS NOT NULL
            """).fetchone()
            n = max(totales["meses_con_datos"], 1)

            canales = [
                {
                    "canal": "Caja",
                    "total": totales["caja"],
                    "promedio_mensual": round(totales["caja"] / n),
                    "pct": round(totales["caja"] / max(totales["total"], 1) * 100, 1),
                },
                {
                    "canal": "Bancos",
                    "total": totales["bancos"],
                    "promedio_mensual": round(totales["bancos"] / n),
                    "pct": round(totales["bancos"] / max(totales["total"], 1) * 100, 1),
                },
                {
                    "canal": "Tarjeta CR",
                    "total": totales["tarjeta_cr"],
                    "promedio_mensual": round(totales["tarjeta_cr"] / n),
                    "pct": round(totales["tarjeta_cr"] / max(totales["total"], 1) * 100, 1),
                },
            ]
            canales.sort(key=lambda c: c["total"], reverse=True)
            return {
                "canales": canales,
                "meses": meses,
                "total_global": totales["total"],
                "meses_con_datos": totales["meses_con_datos"],
            }
        finally:
            conn.close()

    return _api_ok(_cached_runtime_query(("ingresos_analisis",), _load_analisis, ttl_seconds=30))


@app.get("/api/ingresos/{ingreso_id}")
def ingreso_detail(ingreso_id: int):
    return _api_ok(_require_existing("ingresos", ingreso_id, "Ingreso"))


@app.post("/api/ingresos")
def create_ingreso(payload: IngresoPayload):
    try:
        save_ingreso(payload.model_dump())
        return _api_ok(message="Ingreso creado.")
    except Exception as exc:
        _handle_validation(exc)


@app.put("/api/ingresos/{ingreso_id}")
def update_ingreso(ingreso_id: int, payload: IngresoPayload):
    _require_existing("ingresos", ingreso_id, "Ingreso")
    try:
        save_ingreso(payload.model_dump(), ingreso_id=ingreso_id)
        return _api_ok(message="Ingreso actualizado.")
    except Exception as exc:
        _handle_validation(exc)


@app.delete("/api/ingresos/{ingreso_id}")
def remove_ingreso(ingreso_id: int):
    _require_existing("ingresos", ingreso_id, "Ingreso")
    delete_ingreso(ingreso_id)
    return _api_ok(message="Ingreso eliminado.")


# ── Cuadre de Caja ────────────────────────────────────────────────────────────

@app.get("/api/caja/hoy")
def caja_hoy():
    def _load_caja_hoy():
        from datetime import date as _date
        today = _date.today().strftime("%Y-%m-%d")
        cuadre = get_cuadre_caja_by_fecha(today)
        snapshot = get_caja_snapshot_by_fecha(today)
        detalle = get_caja_movimientos_detalle(today)
        movs = detalle["resumen"]
        saldo_sugerido = get_saldo_inicial_sugerido(today)
        apertura = get_caja_apertura_context(today)
        saldo_operativo = float(snapshot["saldo_inicial"]) if snapshot else float(saldo_sugerido or 0)
        saldo_actual = float(snapshot["saldo_esperado"]) if snapshot else round(
            saldo_operativo + float(movs["ingresos_caja"] or 0) - float(movs["egresos_caja"] or 0),
            2,
        )
        saldo_contado = snapshot.get("saldo_real") if snapshot else None
        diferencia_arqueo = round(float(saldo_contado) - saldo_actual, 2) if saldo_contado is not None else None
        return {
            "fecha": today,
            "cuadre": cuadre,
            "snapshot": snapshot,
            "movimientos": movs,
            "saldo_inicial_sugerido": saldo_sugerido,
            "saldo_inicial_operativo": saldo_operativo,
            "saldo_actual": saldo_actual,
            "saldo_contado": saldo_contado,
            "diferencia_arqueo": diferencia_arqueo,
            "detalle_movimientos": detalle,
            "apertura": apertura,
        }

    return _api_ok(
        _cached_runtime_query(
            ("caja_hoy",),
            _load_caja_hoy,
            ttl_seconds=10,
        )
    )


@app.get("/api/caja")
def cuadres_caja(
    mes: int | None = Query(default=None),
    ano: int | None = Query(default=None),
):
    _validate_period(mes, ano)
    return _api_ok(
        _cached_runtime_query(
            ("caja", mes, ano),
            lambda: get_cuadres_caja(mes=mes, ano=ano),
        )
    )


@app.get("/api/caja/ajustes")
def caja_ajustes(
    mes: int | None = Query(default=None),
    ano: int | None = Query(default=None),
):
    _validate_period(mes, ano)
    return _api_ok(
        _cached_runtime_query(
            ("caja_ajustes", mes, ano),
            lambda: get_caja_ajustes(mes=mes, ano=ano),
        )
    )


@app.get("/api/caja/{cuadre_id}")
def cuadre_caja_detail(cuadre_id: int):
    return _api_ok(_require_existing("cuadre_caja", cuadre_id, "Cuadre"))


@app.post("/api/caja")
def create_cuadre_caja(payload: CuadreCajaPayload):
    try:
        cuadre_id = save_cuadre_caja(payload.model_dump())
        return _api_ok({"id": cuadre_id}, "Caja actualizada.")
    except Exception as exc:
        _handle_validation(exc)


@app.put("/api/caja/{cuadre_id}")
def update_cuadre_caja(cuadre_id: int, payload: CuadreCajaPayload):
    _require_existing("cuadre_caja", cuadre_id, "Cuadre")
    try:
        save_cuadre_caja(payload.model_dump(), cuadre_id=cuadre_id)
        return _api_ok(message="Caja actualizada.")
    except Exception as exc:
        _handle_validation(exc)


@app.delete("/api/caja/{cuadre_id}")
def remove_cuadre_caja(cuadre_id: int):
    _require_existing("cuadre_caja", cuadre_id, "Cuadre")
    try:
        delete_cuadre_caja(cuadre_id)
        return _api_ok(message="Cuadre eliminado.")
    except Exception as exc:
        _handle_validation(exc)


@app.post("/api/caja/ajustes")
def create_caja_ajuste_endpoint(payload: CajaAjustePayload):
    try:
        ajuste_id = create_caja_ajuste(payload.model_dump())
        return _api_ok({"id": ajuste_id}, "Ajuste manual registrado.")
    except Exception as exc:
        _handle_validation(exc)


@app.get("/api/egresos")
def egresos(
    mes: int | None = Query(default=None),
    ano: int | None = Query(default=None),
    tipo: str | None = Query(default=None),
    search: str = "",
):
    _validate_period(mes, ano)
    normalized_search = search.strip().lower()
    rows = _cached_runtime_query(
        ("egresos", mes, ano, tipo or "", normalized_search),
        lambda: get_egresos(mes=mes, ano=ano, tipo=tipo, search=search),
    )
    return _api_ok([_serialize_egreso(row) for row in rows])


@app.get("/api/egresos/{egreso_id}")
def egreso_detail(egreso_id: int):
    return _api_ok(_serialize_egreso(_require_existing("egresos", egreso_id, "Egreso")))


@app.get("/api/egresos-meta")
def egresos_meta():
    return _api_ok(
        _cached_runtime_query(
            ("egresos_meta",),
            lambda: {"tipos_gasto": get_tipos_gasto_distintos()},
            ttl_seconds=60,
        )
    )


@app.post("/api/egresos")
def create_egreso(payload: EgresoPayload):
    try:
        egreso_id = save_egreso(payload.model_dump())
        return _api_ok({"id": egreso_id}, "Egreso creado.")
    except Exception as exc:
        _handle_validation(exc)


@app.put("/api/egresos/{egreso_id}")
def update_egreso(egreso_id: int, payload: EgresoPayload):
    _require_existing("egresos", egreso_id, "Egreso")
    try:
        save_egreso(payload.model_dump(), egreso_id=egreso_id)
        return _api_ok(message="Egreso actualizado.")
    except Exception as exc:
        _handle_validation(exc)


@app.delete("/api/egresos/{egreso_id}")
def remove_egreso(egreso_id: int):
    _require_existing("egresos", egreso_id, "Egreso")
    delete_egreso(egreso_id)
    return _api_ok(message="Egreso eliminado.")


@app.post("/api/egresos/{egreso_id}/soporte")
async def upload_egreso_soporte(egreso_id: int, file: UploadFile = File(...)):
    row = _require_existing("egresos", egreso_id, "Egreso")
    previous_file_id = row.get("support_file_id")
    content, filename, total_size, content_type = await _read_uploaded_binary(
        file,
        allowed_extensions=ALLOWED_SUPPORT_EXTENSIONS,
        allowed_content_types=ALLOWED_SUPPORT_CONTENT_TYPES,
        max_size_bytes=MAX_SUPPORT_SIZE_BYTES,
    )
    mes, ano = month_year_from_date(row["fecha"])
    file_id = create_archivo_blob(f"egreso:{egreso_id}", filename, content_type, total_size, content)
    updated = dict(row)
    updated["soporte_path"] = ""
    updated["soporte_name"] = filename
    updated["support_file_id"] = file_id
    try:
        save_egreso(updated, egreso_id=egreso_id)
        if previous_file_id and previous_file_id != file_id:
            delete_archivo_blob(previous_file_id)
        log_auditoria("egreso_soporte", "UPLOAD", egreso_id, period_from_month_year(mes, ano), row["razon_social"], {
            "soporte_name": filename,
            "size_bytes": total_size,
        })
        return _api_ok({"name": filename, "has_support": True}, "Soporte cargado correctamente.")
    except Exception as exc:
        delete_archivo_blob(file_id)
        _handle_validation(exc)


@app.get("/api/egresos/{egreso_id}/soporte")
def get_egreso_soporte(egreso_id: int):
    row = _require_existing("egresos", egreso_id, "Egreso")
    support_file_id = row.get("support_file_id")
    if not support_file_id:
        raise HTTPException(status_code=404, detail="Este egreso no tiene soporte cargado.")
    support_file = get_archivo_blob(support_file_id)
    if not support_file or not support_file.get("content"):
        raise HTTPException(status_code=404, detail="Este egreso no tiene soporte disponible.")
    filename = _sanitize_filename(row.get("soporte_name") or support_file.get("file_name") or f"egreso_{egreso_id}")
    return StreamingResponse(
        BytesIO(bytes(support_file["content"])),
        media_type=support_file.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@app.get("/api/nomina")
def nomina(periodo: str | None = Query(default=None), search: str = ""):
    normalized_search = search.strip().lower()
    return _api_ok(
        _cached_runtime_query(
            ("nomina", periodo or "", normalized_search),
            lambda: get_nomina_bundle(periodo=periodo, search=search),
        )
    )


@app.post("/api/nomina/asistencia")
def create_nomina_asistencia(payload: AsistenciaPayload):
    try:
        save_nomina_asistencia(payload.model_dump())
        return _api_ok(message="Asistencia registrada.")
    except Exception as exc:
        _handle_validation(exc)


@app.put("/api/nomina/asistencia/{asistencia_id}")
def update_nomina_asistencia(asistencia_id: int, payload: AsistenciaPayload):
    _require_existing("nomina_asistencia", asistencia_id, "Asistencia")
    try:
        save_nomina_asistencia(payload.model_dump(), asistencia_id=asistencia_id)
        return _api_ok(message="Asistencia actualizada.")
    except Exception as exc:
        _handle_validation(exc)


@app.delete("/api/nomina/asistencia/{asistencia_id}")
def remove_nomina_asistencia(asistencia_id: int):
    _require_existing("nomina_asistencia", asistencia_id, "Asistencia")
    delete_nomina_asistencia(asistencia_id)
    return _api_ok(message="Asistencia eliminada.")


@app.post("/api/nomina/novedades")
def create_nomina_novedad(payload: NovedadPayload):
    try:
        save_nomina_novedad(payload.model_dump())
        return _api_ok(message="Novedad creada.")
    except Exception as exc:
        _handle_validation(exc)


@app.put("/api/nomina/novedades/{novedad_id}")
def update_nomina_novedad(novedad_id: int, payload: NovedadPayload):
    _require_existing("nomina_novedades", novedad_id, "Novedad")
    try:
        save_nomina_novedad(payload.model_dump(), novedad_id=novedad_id)
        return _api_ok(message="Novedad actualizada.")
    except Exception as exc:
        _handle_validation(exc)


@app.delete("/api/nomina/novedades/{novedad_id}")
def remove_nomina_novedad(novedad_id: int):
    _require_existing("nomina_novedades", novedad_id, "Novedad")
    delete_nomina_novedad(novedad_id)
    return _api_ok(message="Novedad eliminada.")


@app.post("/api/nomina/sync")
def sync_nomina(payload: SyncNominaPayload):
    total = sync_nomina_to_egresos(periodo=payload.periodo)
    getattr(app.state, "runtime", _startup_state())["db_health"] = _current_db_health()
    return _api_ok({"egresos_generados": total}, "Nómina sincronizada con egresos.")


@app.get("/api/reportes/cierre")
def cierre(mes: int = Query(...), ano: int = Query(...), include_details: bool = Query(default=True)):
    _validate_period(mes, ano)
    def _load_cierre():
        periodo = period_from_month_year(mes, ano)
        cierre_data = get_cierre_mensual(mes, ano)
        if not include_details:
            return {"cierre": cierre_data}
        egresos_rows = get_egresos(mes=mes, ano=ano)
        return {
            "cierre": cierre_data,
            "ingresos": get_ingresos(mes=mes, ano=ano),
            "egresos": [_serialize_egreso(row) for row in egresos_rows],
            "nomina": get_nomina_resumen(periodo=periodo),
            "novedades": get_nomina_novedades(periodo=periodo),
            "seg_social": get_nomina_seg_social(periodo=periodo),
        }

    return _api_ok(
        _cached_runtime_query(
            ("reportes_cierre", mes, ano, include_details),
            _load_cierre,
        )
    )


@app.get("/api/cierres")
def cierres():
    return _api_ok(
        _cached_runtime_query(
            ("cierres",),
            list_cierres_mensuales,
        )
    )


@app.post("/api/cierres/cerrar")
def cerrar_mes(payload: CierreMensualPayload):
    _validate_period(payload.mes, payload.ano)
    cierre_data = get_cierre_mensual(payload.mes, payload.ano)
    if not cierre_data["periodo"]:
        raise HTTPException(status_code=400, detail="No se pudo determinar el período.")
    if (cierre_data.get("total_ingresos") or 0) == 0 and (cierre_data.get("total_egresos") or 0) == 0:
        raise HTTPException(status_code=400, detail="No puedes cerrar un período sin movimientos.")
    result = set_cierre_mensual(payload.mes, payload.ano, True, payload.observacion)
    return _api_ok(result, "Mes cerrado correctamente.")


@app.post("/api/cierres/reabrir")
def reabrir_mes(payload: CierreMensualPayload):
    _validate_period(payload.mes, payload.ano)
    result = set_cierre_mensual(payload.mes, payload.ano, False, payload.observacion)
    return _api_ok(result, "Mes reabierto correctamente.")


@app.get("/api/auditoria")
def auditoria(limit: int = Query(default=120, ge=1, le=500)):
    return _api_ok(
        _cached_runtime_query(
            ("auditoria", limit),
            lambda: [_serialize_audit_entry(row) for row in get_auditoria(limit=limit)],
            ttl_seconds=15,
        )
    )


@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...)):
    source, source_name, total_size = await _save_import_upload(file, "excel")
    try:
        migrate_excel_module.migrate(path=str(source))
        getattr(app.state, "runtime", _startup_state())["db_health"] = _current_db_health()
        return _api_ok(
            {"source_name": source_name, "size_bytes": total_size, "counts": _system_counts()},
            "Excel importado correctamente.",
        )
    except SystemExit as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo importar el Excel: {exc}") from exc
    except Exception as exc:
        _handle_validation(exc)
    finally:
        source.unlink(missing_ok=True)


@app.post("/api/import/nomina")
async def import_nomina(file: UploadFile = File(...)):
    source, source_name, total_size = await _save_import_upload(file, "nomina")
    try:
        migrate_nomina_module.migrate_nomina(path=str(source))
        generated = sync_nomina_to_egresos()
        getattr(app.state, "runtime", _startup_state())["db_health"] = _current_db_health()
        return _api_ok(
            {
                "source_name": source_name,
                "size_bytes": total_size,
                "egresos_generados": generated,
                "counts": _system_counts(),
            },
            "Nómina importada correctamente.",
        )
    except SystemExit as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo importar la nómina: {exc}") from exc
    except Exception as exc:
        _handle_validation(exc)
    finally:
        source.unlink(missing_ok=True)


def _build_excel_response(workbook, filename: str):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_export_workbook(sheet_name: str, headers, rows, title: str | None = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    current_row = 1
    if title:
        ws.cell(current_row, 1, _sanitize_excel_value(title))
        ws.cell(current_row, 1).font = Font(bold=True, size=14)
        current_row += 2

    header_fill = PatternFill('solid', fgColor='1e3a5f')
    white_font = Font(color='FFFFFF', bold=True)
    center = Alignment(horizontal='center')

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(current_row, col, _sanitize_excel_value(label))
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center

    for row_index, row in enumerate(rows, start=current_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(row_index, col, _sanitize_excel_value(value))

    return wb


@app.get("/api/export/reportes")
def export_reportes(mes: int = Query(...), ano: int = Query(...)):
    _validate_period(mes, ano)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    cierre = get_cierre_mensual(mes, ano)
    periodo_nomina = period_from_month_year(mes, ano)
    egresos = get_egresos(mes=mes, ano=ano)
    ingresos = get_ingresos(mes=mes, ano=ano)
    nomina = get_nomina_resumen(periodo=periodo_nomina)
    novedades = get_nomina_novedades(periodo=periodo_nomina)
    seg = get_nomina_seg_social(periodo=periodo_nomina)

    wb = openpyxl.Workbook()
    header_fill = PatternFill('solid', fgColor='1e3a5f')
    green_fill = PatternFill('solid', fgColor='16a34a')
    red_fill = PatternFill('solid', fgColor='dc2626')
    blue_fill = PatternFill('solid', fgColor='1d4ed8')
    purple_fill = PatternFill('solid', fgColor='7c3aed')
    white_font = Font(color='FFFFFF', bold=True)
    center = Alignment(horizontal='center')

    ws = wb.active
    ws.title = 'Cierre'
    ws['A1'] = _sanitize_excel_value(f'CIERRE MENSUAL {mes:02d}-{ano}')
    ws['A1'].font = Font(bold=True, size=14)
    for i, (label, value) in enumerate([
        ('Ingresos', cierre['total_ingresos']),
        ('Egresos operativos', cierre['egresos_operativos']),
        ('Nomina empleados', cierre['egresos_nomina']),
        ('Seguridad social', cierre['egresos_seg_social']),
        ('Novedades deduccion', cierre['novedades_deduccion']),
        ('Total egresos', cierre['total_egresos']),
        ('Resultado neto', cierre['resultado_neto']),
    ], start=3):
        ws.cell(i, 1, _sanitize_excel_value(label))
        ws.cell(i, 2, _sanitize_excel_value(value)).number_format = '$ #,##0'

    def make_sheet(name, headers, rows, fill):
      sheet = wb.create_sheet(name)
      for c, label in enumerate(headers, 1):
          cell = sheet.cell(1, c, _sanitize_excel_value(label))
          cell.fill = fill
          cell.font = white_font
          cell.alignment = center
      for r, row in enumerate(rows, 2):
          for c, value in enumerate(row, 1):
              sheet.cell(r, c, _sanitize_excel_value(value))

    make_sheet(
        'Egresos',
        ['Fecha', 'Proveedor', 'Naturaleza', 'Factura', 'Valor'],
        [(r['fecha'], r['razon_social'], r['tipo_gasto'], r.get('factura_electronica', 'NO'), r['valor']) for r in egresos],
        red_fill,
    )
    make_sheet(
        'Ingresos',
        ['Fecha', 'Caja', 'Bancos', 'Tarjeta'],
        [(r['fecha'], r['caja'], r['bancos'], r['tarjeta_cr']) for r in ingresos],
        green_fill,
    )
    make_sheet(
        'Nomina',
        ['Empleado', 'Cedula', 'Neto Q1', 'Neto Q2', 'Total Mes'],
        [(r['empleado'], r['cedula'], r['q1_neto'], r['q2_neto'], r['total_mes']) for r in nomina],
        blue_fill,
    )
    make_sheet(
        'Novedades',
        ['Fecha', 'Empleado', 'Naturaleza', 'Tipo', 'Valor'],
        [(r['fecha'], r['empleado'], r['naturaleza'], r['tipo_novedad'], r['valor']) for r in novedades],
        purple_fill,
    )
    make_sheet(
        'SegSocial',
        ['Grupo', 'Concepto', 'Valor'],
        [(r['grupo'], r['concepto'], r['valor']) for r in seg],
        purple_fill,
    )
    return _build_excel_response(wb, f'Reporte_{mes:02d}_{ano}.xlsx')


@app.get("/api/export/proveedores")
def export_proveedores(search: str = ""):
    proveedores = get_proveedores(search=search)
    wb = _build_export_workbook(
        'Proveedores',
        ['Razon Social', 'NIT', 'Telefono', 'Correo', 'Direccion'],
        [
            (
                row.get('razon_social', ''),
                row.get('nit', ''),
                row.get('telefono', ''),
                row.get('correo', ''),
                row.get('direccion', ''),
            )
            for row in proveedores
        ],
        title='BASE DE PROVEEDORES',
    )
    return _build_excel_response(wb, 'Proveedores.xlsx')


@app.get("/api/export/ingresos")
def export_ingresos(mes: int = Query(...), ano: int = Query(...)):
    _validate_period(mes, ano)
    ingresos = get_ingresos(mes=mes, ano=ano)
    wb = _build_export_workbook(
        'Ingresos',
        ['Fecha', 'Caja', 'Bancos', 'Tarjeta CR', 'Total Dia'],
        [
            (
                row.get('fecha', ''),
                row.get('caja', 0),
                row.get('bancos', 0),
                row.get('tarjeta_cr', 0),
                (row.get('caja') or 0) + (row.get('bancos') or 0) + (row.get('tarjeta_cr') or 0),
            )
            for row in ingresos
        ],
        title=f'INGRESOS {mes:02d}-{ano}',
    )
    return _build_excel_response(wb, f'Ingresos_{mes:02d}_{ano}.xlsx')


@app.get("/api/export/egresos")
def export_egresos(
    mes: int = Query(...),
    ano: int = Query(...),
    tipo: str | None = Query(default=None),
    search: str = "",
):
    _validate_period(mes, ano)
    egresos = get_egresos(mes=mes, ano=ano, tipo=tipo, search=search)
    wb = _build_export_workbook(
        'Egresos',
        ['Fecha', 'N Documento', 'Proveedor', 'NIT', 'Naturaleza del gasto', 'Factura electronica', 'Valor', 'Observaciones'],
        [
            (
                row.get('fecha', ''),
                row.get('no_documento', ''),
                row.get('razon_social', ''),
                row.get('nit', ''),
                row.get('tipo_gasto', ''),
                row.get('factura_electronica', 'NO'),
                row.get('valor', 0),
                row.get('observaciones', ''),
            )
            for row in egresos
        ],
        title=f'EGRESOS {mes:02d}-{ano}',
    )
    return _build_excel_response(wb, f'Egresos_{mes:02d}_{ano}.xlsx')


@app.get("/api/export/nomina")
def export_nomina(periodo: str = Query(...)):
    resumen = get_nomina_resumen(periodo=periodo)
    novedades = get_nomina_novedades(periodo=periodo)
    seg = get_nomina_seg_social(periodo=periodo)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    header_fill = PatternFill('solid', fgColor='1e3a5f')
    white_font = Font(color='FFFFFF', bold=True)
    center = Alignment(horizontal='center')

    def make_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(1, col, _sanitize_excel_value(label))
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
        for row_index, row in enumerate(rows, start=2):
            for col, value in enumerate(row, start=1):
                ws.cell(row_index, col, _sanitize_excel_value(value))

    ws = wb.active
    ws.title = 'Resumen'
    ws['A1'] = _sanitize_excel_value(f'NOMINA {periodo}')
    ws['A1'].font = Font(bold=True, size=14)

    make_sheet(
        'Liquidacion',
        ['Empleado', 'Cedula', 'Valor dia', 'Dias Q1', 'Neto Q1', 'Dias Q2', 'Neto Q2', 'Deducciones', 'Total mes'],
        [
            (
                row.get('empleado', ''),
                row.get('cedula', ''),
                row.get('valor_dia', 0),
                row.get('q1_dias', 0),
                row.get('q1_neto', 0),
                row.get('q2_dias', 0),
                row.get('q2_neto', 0),
                row.get('total_deduccion', 0),
                row.get('total_mes', 0),
            )
            for row in resumen
        ],
    )
    make_sheet(
        'Novedades',
        ['Fecha', 'Empleado', 'Quincena', 'Naturaleza', 'Tipo', 'Valor', 'Observaciones'],
        [
            (
                row.get('fecha', ''),
                row.get('empleado', ''),
                row.get('quincena', ''),
                row.get('naturaleza', ''),
                row.get('tipo_novedad', ''),
                row.get('valor', 0),
                row.get('observaciones', ''),
            )
            for row in novedades
        ],
    )
    make_sheet(
        'Seguridad Social',
        ['Grupo', 'Concepto', 'Valor', 'Observaciones'],
        [
            (
                row.get('grupo', ''),
                row.get('concepto', ''),
                row.get('valor', 0),
                row.get('observaciones', ''),
            )
            for row in seg
        ],
    )
    return _build_excel_response(wb, f'Nomina_{periodo.replace(" ", "_")}.xlsx')


if FRONTEND_DIST_DIR.exists():
    assets_dir = FRONTEND_DIST_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="frontend-assets")

    @app.get("/favicon.svg", include_in_schema=False)
    def frontend_favicon():
        favicon = FRONTEND_DIST_DIR / "favicon.svg"
        if favicon.exists():
            return FileResponse(favicon)
        raise HTTPException(status_code=404, detail="Favicon no encontrado.")

    @app.get("/icons.svg", include_in_schema=False)
    def frontend_icons():
        icons = FRONTEND_DIST_DIR / "icons.svg"
        if icons.exists():
            return FileResponse(icons)
        raise HTTPException(status_code=404, detail="Iconos no encontrados.")

    @app.get("/{full_path:path}", include_in_schema=False)
    def frontend_spa(full_path: str):
        if full_path.startswith("api/") or full_path in {"docs", "openapi.json", "health"}:
            raise HTTPException(status_code=404, detail="Not Found")
        return _frontend_index_response()
